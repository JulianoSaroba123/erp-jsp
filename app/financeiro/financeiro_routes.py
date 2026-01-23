# -*- coding: utf-8 -*-
"""
ERP JSP v3.0 - Rotas do Financeiro
=================================

Rotas para gerenciamento financeiro.
Inclui lançamentos, contas a pagar e receber.

Autor: JSP Soluções
Data: 2025
"""

from flask import Blueprint, render_template, redirect, url_for, request, flash, jsonify, send_file
from datetime import datetime, date, timedelta
from decimal import Decimal
import io
from app.extensoes import db
from app.financeiro.financeiro_model import (
    LancamentoFinanceiro, CategoriaFinanceira, ContaBancaria, CentroCusto, CustoFixo,
    Notificacao, RateioDespesa, ImportacaoLote, RelatorioCustomizado
)
from app.cliente.cliente_model import Cliente
from app.fornecedor.fornecedor_model import Fornecedor

# Criar blueprint
bp_financeiro = Blueprint('financeiro', __name__, template_folder='templates')


def converter_valor_monetario(valor_str):
    """Converte valor monetário do formato brasileiro para decimal."""
    if not valor_str:
        return Decimal('0.00')
    
    # Remove R$, espaços e pontos (separadores de milhares)
    valor_limpo = valor_str.replace('R$', '').replace(' ', '').replace('.', '')
    
    # Substitui vírgula por ponto (separador decimal)
    valor_limpo = valor_limpo.replace(',', '.')
    
    try:
        return Decimal(valor_limpo)
    except:
        return Decimal('0.00')


@bp_financeiro.route('/')
def dashboard():
    """Dashboard financeiro com resumo e indicadores."""
    try:
        # Verificar se as tabelas existem antes de consultar
        from sqlalchemy import inspect
        inspector = inspect(db.engine)
        tabelas = inspector.get_table_names()
        
        # Se tabelas não existem, mostrar mensagem amigável
        if 'lancamentos_financeiros' not in tabelas:
            flash('⚠️ Tabelas financeiras não encontradas. Execute o script de criação.', 'warning')
            return render_template('financeiro/dashboard.html',
                                 resumo=None,
                                 vencidos=0,
                                 pendentes=0,
                                 contas_receber=0,
                                 contas_pagar=0,
                                 ultimos_lancamentos=[],
                                 data_atual=date.today())
        
        # Resumo do mês atual
        resumo = LancamentoFinanceiro.get_resumo_mes()
        
        # Lançamentos vencidos
        vencidos = LancamentoFinanceiro.get_vencidos().count()
        
        # Pendentes
        pendentes = LancamentoFinanceiro.get_pendentes().count()
        
        # Contas a receber este mês
        contas_receber = LancamentoFinanceiro.get_contas_receber().filter(
            db.extract('month', LancamentoFinanceiro.data_vencimento) == date.today().month
        ).count()
        
        # Contas a pagar este mês
        contas_pagar = LancamentoFinanceiro.get_contas_pagar().filter(
            db.extract('month', LancamentoFinanceiro.data_vencimento) == date.today().month
        ).count()
        
        # Últimos lançamentos (usando data_criacao_auditoria ou data_lancamento)
        ultimos_lancamentos = LancamentoFinanceiro.query.filter_by(ativo=True).order_by(
            LancamentoFinanceiro.data_lancamento.desc()
        ).limit(10).all()
        
        return render_template('financeiro/dashboard.html',
                             resumo=resumo,
                             vencidos=vencidos,
                             pendentes=pendentes,
                             contas_receber=contas_receber,
                             contas_pagar=contas_pagar,
                             ultimos_lancamentos=ultimos_lancamentos,
                             data_atual=date.today())
    
    except Exception as e:
        print(f"❌ Erro no dashboard financeiro: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Erro ao carregar dashboard: {str(e)}', 'danger')
        return render_template('financeiro/dashboard.html',
                             resumo=None,
                             vencidos=0,
                             pendentes=0,
                             contas_receber=0,
                             contas_pagar=0,
                             ultimos_lancamentos=[],
                             data_atual=date.today())


@bp_financeiro.route('/lancamentos')
def listar_lancamentos():
    """Lista todos os lançamentos financeiros."""
    try:
        print("DEBUG: Iniciando listar_lancamentos")
        
        # Filtros
        tipo = request.args.get('tipo')
        status = request.args.get('status')
        categoria = request.args.get('categoria')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        
        print(f"DEBUG: Filtros: tipo={tipo}, status={status}")
        
        # Query base
        query = LancamentoFinanceiro.query.filter_by(ativo=True)
        
        # Aplicar filtros
        if tipo:
            query = query.filter(LancamentoFinanceiro.tipo == tipo)
        
        if status:
            query = query.filter(LancamentoFinanceiro.status == status)
        
        if categoria:
            query = query.filter(LancamentoFinanceiro.categoria == categoria)
        
        if data_inicio:
            data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            query = query.filter(LancamentoFinanceiro.data_lancamento >= data_inicio_obj)
        
        if data_fim:
            data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
            query = query.filter(LancamentoFinanceiro.data_lancamento <= data_fim_obj)
        
        # Ordenação
        lancamentos = query.order_by(LancamentoFinanceiro.data_lancamento.desc()).all()
        print(f"DEBUG: Encontrados {len(lancamentos)} lançamentos")
        
        # Categorias para filtro
        categorias = db.session.query(LancamentoFinanceiro.categoria).filter(
            LancamentoFinanceiro.categoria.isnot(None),
            LancamentoFinanceiro.ativo == True
        ).distinct().all()
        categorias = [c[0] for c in categorias if c[0]]
        
        print("DEBUG: Renderizando template")
        return render_template('financeiro/listar_lancamentos.html',
                             lancamentos=lancamentos,
                             categorias=categorias,
                             filtros={
                                 'tipo': tipo,
                                 'status': status,
                                 'categoria': categoria,
                                 'data_inicio': data_inicio,
                                 'data_fim': data_fim
                             })
    
    except Exception as e:
        print(f"DEBUG: Erro em listar_lancamentos: {str(e)}")
        flash(f'Erro ao listar lançamentos: {str(e)}', 'danger')
        return redirect(url_for('financeiro.dashboard'))


@bp_financeiro.route('/lancamentos/novo')
def novo_lancamento():
    """Formulário para novo lançamento."""
    try:
        clientes = Cliente.query.filter_by(ativo=True).all()
        fornecedores = Fornecedor.query.filter_by(ativo=True).all()
        contas_bancarias = ContaBancaria.query.filter_by(ativo=True, ativa=True).all()
        centros_custo = CentroCusto.query.filter_by(ativo=True).all()
        
        return render_template('financeiro/form_lancamento.html',
                             clientes=clientes,
                             fornecedores=fornecedores,
                             contas_bancarias=contas_bancarias,
                             centros_custo=centros_custo,
                             date=date)
    
    except Exception as e:
        flash(f'Erro ao carregar formulário: {str(e)}', 'danger')
        return redirect(url_for('financeiro.listar_lancamentos'))


@bp_financeiro.route('/lancamentos/criar', methods=['POST'])
def criar_lancamento():
    """Cria novo lançamento financeiro."""
    try:
        # Dados do formulário
        descricao = request.form.get('descricao')
        valor_str = request.form.get('valor')
        tipo = request.form.get('tipo')
        categoria = request.form.get('categoria')
        subcategoria = request.form.get('subcategoria')
        data_lancamento_str = request.form.get('data_lancamento')
        data_vencimento_str = request.form.get('data_vencimento')
        status = request.form.get('status', 'pendente')
        observacoes = request.form.get('observacoes')
        numero_documento = request.form.get('numero_documento')
        forma_pagamento = request.form.get('forma_pagamento')
        cliente_id = request.form.get('cliente_id')
        fornecedor_id = request.form.get('fornecedor_id')
        conta_bancaria_id = request.form.get('conta_bancaria_id')
        centro_custo_id = request.form.get('centro_custo_id')
        recorrente = bool(request.form.get('recorrente'))
        frequencia = request.form.get('frequencia')
        
        # Validações básicas
        if not descricao:
            flash('Descrição é obrigatória', 'danger')
            return redirect(url_for('financeiro.novo_lancamento'))
        
        if not valor_str:
            flash('Valor é obrigatório', 'danger')
            return redirect(url_for('financeiro.novo_lancamento'))
        
        if not tipo:
            flash('Tipo é obrigatório', 'danger')
            return redirect(url_for('financeiro.novo_lancamento'))
        
        # Conversões
        valor = converter_valor_monetario(valor_str)
        
        if valor <= 0:
            flash('Valor deve ser maior que zero', 'danger')
            return redirect(url_for('financeiro.novo_lancamento'))
        
        # Datas
        data_lancamento = datetime.strptime(data_lancamento_str, '%Y-%m-%d').date() if data_lancamento_str else date.today()
        data_vencimento = datetime.strptime(data_vencimento_str, '%Y-%m-%d').date() if data_vencimento_str else None
        
        # Criar lançamento
        lancamento = LancamentoFinanceiro(
            descricao=descricao,
            valor=valor,
            tipo=tipo,
            categoria=categoria,
            subcategoria=subcategoria,
            data_lancamento=data_lancamento,
            data_vencimento=data_vencimento,
            status=status,
            observacoes=observacoes,
            numero_documento=numero_documento,
            forma_pagamento=forma_pagamento,
            cliente_id=int(cliente_id) if cliente_id else None,
            fornecedor_id=int(fornecedor_id) if fornecedor_id else None,
            conta_bancaria_id=int(conta_bancaria_id) if conta_bancaria_id else None,
            centro_custo_id=int(centro_custo_id) if centro_custo_id else None,
            recorrente=recorrente,
            frequencia=frequencia if recorrente else None
        )
        
        db.session.add(lancamento)
        db.session.commit()
        
        flash(f'Lançamento "{descricao}" criado com sucesso!', 'success')
        return redirect(url_for('financeiro.listar_lancamentos'))
    
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao criar lançamento: {str(e)}', 'danger')
        return redirect(url_for('financeiro.novo_lancamento'))


@bp_financeiro.route('/lancamentos/<int:id>/editar')
def editar_lancamento(id):
    """Formulário para editar lançamento."""
    try:
        lancamento = LancamentoFinanceiro.query.get_or_404(id)
        clientes = Cliente.query.filter_by(ativo=True).all()
        fornecedores = Fornecedor.query.filter_by(ativo=True).all()
        contas_bancarias = ContaBancaria.query.filter_by(ativo=True, ativa=True).all()
        centros_custo = CentroCusto.query.filter_by(ativo=True).all()
        
        return render_template('financeiro/form_lancamento.html',
                             lancamento=lancamento,
                             clientes=clientes,
                             fornecedores=fornecedores,
                             contas_bancarias=contas_bancarias,
                             centros_custo=centros_custo,
                             date=date)
    
    except Exception as e:
        flash(f'Erro ao carregar lançamento: {str(e)}', 'danger')
        return redirect(url_for('financeiro.listar_lancamentos'))


@bp_financeiro.route('/lancamentos/<int:id>/atualizar', methods=['POST'])
def atualizar_lancamento(id):
    """Atualiza lançamento financeiro."""
    try:
        lancamento = LancamentoFinanceiro.query.get_or_404(id)
        
        # Dados do formulário
        lancamento.descricao = request.form.get('descricao')
        valor_str = request.form.get('valor')
        lancamento.tipo = request.form.get('tipo')
        lancamento.categoria = request.form.get('categoria')
        lancamento.subcategoria = request.form.get('subcategoria')
        data_lancamento_str = request.form.get('data_lancamento')
        data_vencimento_str = request.form.get('data_vencimento')
        lancamento.status = request.form.get('status', 'pendente')
        lancamento.observacoes = request.form.get('observacoes')
        lancamento.numero_documento = request.form.get('numero_documento')
        lancamento.forma_pagamento = request.form.get('forma_pagamento')
        cliente_id = request.form.get('cliente_id')
        fornecedor_id = request.form.get('fornecedor_id')
        conta_bancaria_id = request.form.get('conta_bancaria_id')
        centro_custo_id = request.form.get('centro_custo_id')
        lancamento.recorrente = bool(request.form.get('recorrente'))
        lancamento.frequencia = request.form.get('frequencia')
        
        # Validações
        if not lancamento.descricao:
            flash('Descrição é obrigatória', 'danger')
            return redirect(url_for('financeiro.editar_lancamento', id=id))
        
        # Conversões
        lancamento.valor = converter_valor_monetario(valor_str)
        
        if lancamento.valor <= 0:
            flash('Valor deve ser maior que zero', 'danger')
            return redirect(url_for('financeiro.editar_lancamento', id=id))
        
        # Datas
        if data_lancamento_str:
            lancamento.data_lancamento = datetime.strptime(data_lancamento_str, '%Y-%m-%d').date()
        
        if data_vencimento_str:
            lancamento.data_vencimento = datetime.strptime(data_vencimento_str, '%Y-%m-%d').date()
        else:
            lancamento.data_vencimento = None
        
        # Relacionamentos
        lancamento.cliente_id = int(cliente_id) if cliente_id else None
        lancamento.fornecedor_id = int(fornecedor_id) if fornecedor_id else None
        lancamento.conta_bancaria_id = int(conta_bancaria_id) if conta_bancaria_id else None
        lancamento.centro_custo_id = int(centro_custo_id) if centro_custo_id else None
        
        if not lancamento.recorrente:
            lancamento.frequencia = None
        
        db.session.commit()
        
        flash(f'Lançamento "{lancamento.descricao}" atualizado com sucesso!', 'success')
        return redirect(url_for('financeiro.listar_lancamentos'))
    
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar lançamento: {str(e)}', 'danger')
        return redirect(url_for('financeiro.editar_lancamento', id=id))


@bp_financeiro.route('/lancamentos/<int:id>/excluir', methods=['GET', 'POST'])
def excluir_lancamento(id):
    """Exclui lançamento financeiro (soft delete)."""
    try:
        lancamento = LancamentoFinanceiro.query.get_or_404(id)
        
        # Se for GET, mostrar página de confirmação
        if request.method == 'GET':
            return render_template('financeiro/confirmar_exclusao.html', 
                                 lancamento=lancamento,
                                 data_hoje=datetime.now().date())
        
        # Se for POST, executar exclusão
        if request.method == 'POST':
            # Soft delete
            lancamento.ativo = False
            lancamento.data_exclusao = datetime.utcnow()
            
            db.session.commit()
            
            flash(f'Lançamento "{lancamento.descricao}" excluído com sucesso!', 'success')
            return redirect(url_for('financeiro.listar_lancamentos'))
    
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir lançamento: {str(e)}', 'danger')
        return redirect(url_for('financeiro.listar_lancamentos'))


@bp_financeiro.route('/lancamentos/<int:id>/pagar', methods=['POST'])
def marcar_como_pago(id):
    """Marca lançamento como pago/recebido."""
    try:
        lancamento = LancamentoFinanceiro.query.get_or_404(id)
        data_pagamento_str = request.form.get('data_pagamento')
        
        if data_pagamento_str:
            data_pagamento = datetime.strptime(data_pagamento_str, '%Y-%m-%d').date()
        else:
            data_pagamento = date.today()
        
        lancamento.marcar_como_pago(data_pagamento)
        
        action = 'pago' if lancamento.tipo in ['despesa', 'conta_pagar'] else 'recebido'
        flash(f'Lançamento marcado como {action}!', 'success')
        
        return redirect(url_for('financeiro.listar_lancamentos'))
    
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao marcar como pago: {str(e)}', 'danger')
        return redirect(url_for('financeiro.listar_lancamentos'))


@bp_financeiro.route('/contas-pagar')
def contas_pagar():
    """Lista contas a pagar."""
    try:
        contas = LancamentoFinanceiro.get_contas_pagar().order_by(
            LancamentoFinanceiro.data_vencimento.asc()
        ).all()
        
        # Estatísticas
        total_pendente = sum(float(c.valor) for c in contas if c.status == 'pendente')
        vencidas = [c for c in contas if c.situacao_vencimento == 'vencido' and c.status == 'pendente']
        
        return render_template('financeiro/contas_pagar.html',
                             contas=contas,
                             total_pendente=total_pendente,
                             vencidas=vencidas,
                             date=date)
    
    except Exception as e:
        flash(f'Erro ao carregar contas a pagar: {str(e)}', 'danger')
        return redirect(url_for('financeiro.dashboard'))


@bp_financeiro.route('/contas-receber')
def contas_receber():
    """Lista contas a receber."""
    try:
        contas = LancamentoFinanceiro.get_contas_receber().order_by(
            LancamentoFinanceiro.data_vencimento.asc()
        ).all()
        
        # Estatísticas
        total_pendente = sum(float(c.valor) for c in contas if c.status == 'pendente')
        vencidas = [c for c in contas if c.situacao_vencimento == 'vencido' and c.status == 'pendente']
        
        return render_template('financeiro/contas_receber.html',
                             contas=contas,
                             total_pendente=total_pendente,
                             vencidas=vencidas,
                             date=date)
    
    except Exception as e:
        flash(f'Erro ao carregar contas a receber: {str(e)}', 'danger')
        return redirect(url_for('financeiro.dashboard'))


# API Routes
@bp_financeiro.route('/api/resumo-mes')
def api_resumo_mes():
    """API para resumo financeiro do mês."""
    try:
        mes = request.args.get('mes', type=int)
        ano = request.args.get('ano', type=int)
        
        resumo = LancamentoFinanceiro.get_resumo_mes(mes, ano)
        
        return jsonify({
            'status': 'success',
            'data': resumo
        })
    
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@bp_financeiro.route('/chaves-documentos')
def chaves_documentos():
    """Página de chaves de documentos fiscais e tributários."""
    return render_template('financeiro/chaves_documentos.html')


@bp_financeiro.route('/api/indicadores')
def api_indicadores():
    """API para indicadores financeiros."""
    try:
        # Lançamentos vencidos
        vencidos = LancamentoFinanceiro.get_vencidos().count()
        
        # Pendentes
        pendentes = LancamentoFinanceiro.get_pendentes().count()
        
        # Resumo do mês
        resumo = LancamentoFinanceiro.get_resumo_mes()
        
        return jsonify({
            'status': 'success',
            'data': {
                'vencidos': vencidos,
                'pendentes': pendentes,
                'resumo_mes': resumo
            }
        })
    
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@bp_financeiro.route('/api/dashboard-dados')
def api_dashboard_dados():
    """API completa para dados do dashboard com gráficos."""
    try:
        from datetime import timedelta
        from sqlalchemy import func, extract
        
        hoje = date.today()
        mes_atual = hoje.month
        ano_atual = hoje.year
        
        # Resumo do mês atual
        resumo_mes = LancamentoFinanceiro.get_resumo_mes()
        
        # Evolução dos últimos 6 meses
        meses_labels = []
        receitas_mes = []
        despesas_mes = []
        saldos_acumulados = []
        saldo_acumulado = 0
        
        for i in range(5, -1, -1):
            mes_ref = hoje - timedelta(days=i*30)
            mes = mes_ref.month
            ano = mes_ref.year
            
            # Nome do mês
            meses_nomes = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 
                          'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
            meses_labels.append(meses_nomes[mes - 1])
            
            # Receitas do mês
            receitas = db.session.query(func.sum(LancamentoFinanceiro.valor)).filter(
                extract('month', LancamentoFinanceiro.data_lancamento) == mes,
                extract('year', LancamentoFinanceiro.data_lancamento) == ano,
                LancamentoFinanceiro.tipo.in_(['receita', 'conta_receber']),
                LancamentoFinanceiro.status == 'recebido',
                LancamentoFinanceiro.ativo == True
            ).scalar() or 0
            
            # Despesas do mês
            despesas = db.session.query(func.sum(LancamentoFinanceiro.valor)).filter(
                extract('month', LancamentoFinanceiro.data_lancamento) == mes,
                extract('year', LancamentoFinanceiro.data_lancamento) == ano,
                LancamentoFinanceiro.tipo.in_(['despesa', 'conta_pagar']),
                LancamentoFinanceiro.status == 'pago',
                LancamentoFinanceiro.ativo == True
            ).scalar() or 0
            
            receitas_mes.append(float(receitas))
            despesas_mes.append(float(despesas))
            
            # Saldo acumulado
            saldo_acumulado += float(receitas) - float(despesas)
            saldos_acumulados.append(saldo_acumulado)
        
        # Top 5 categorias de despesas
        categorias_despesas = db.session.query(
            LancamentoFinanceiro.categoria,
            func.sum(LancamentoFinanceiro.valor).label('total')
        ).filter(
            extract('month', LancamentoFinanceiro.data_lancamento) == mes_atual,
            extract('year', LancamentoFinanceiro.data_lancamento) == ano_atual,
            LancamentoFinanceiro.tipo.in_(['despesa', 'conta_pagar']),
            LancamentoFinanceiro.categoria.isnot(None),
            LancamentoFinanceiro.ativo == True
        ).group_by(
            LancamentoFinanceiro.categoria
        ).order_by(
            func.sum(LancamentoFinanceiro.valor).desc()
        ).limit(5).all()
        
        categorias_nomes = [cat[0] or 'Sem categoria' for cat in categorias_despesas]
        categorias_valores = [float(cat[1]) for cat in categorias_despesas]
        
        # Cores para categorias
        cores_categorias = [
            'rgba(220, 53, 69, 0.7)',
            'rgba(255, 193, 7, 0.7)',
            'rgba(23, 162, 184, 0.7)',
            'rgba(108, 117, 125, 0.7)',
            'rgba(0, 123, 255, 0.7)'
        ]
        
        # Fluxo de caixa projetado (30 dias)
        fluxo_projetado = LancamentoFinanceiro.calcular_fluxo_caixa(30)
        
        return jsonify({
            'status': 'success',
            'data': {
                'resumo_mes': resumo_mes,
                'evolucao_mensal': {
                    'meses': meses_labels,
                    'receitas': receitas_mes,
                    'despesas': despesas_mes
                },
                'fluxo_caixa': {
                    'meses': meses_labels,
                    'saldos': saldos_acumulados
                },
                'top_categorias': {
                    'categorias': categorias_nomes,
                    'valores': categorias_valores,
                    'cores': cores_categorias[:len(categorias_nomes)]
                },
                'fluxo_projetado': fluxo_projetado
            }
        })
    
    except Exception as e:
        print(f"Erro na API dashboard: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


# ========================================
# CONTAS BANCÁRIAS
# ========================================

@bp_financeiro.route('/contas-bancarias')
def listar_contas_bancarias():
    """Lista todas as contas bancárias."""
    try:
        contas = ContaBancaria.query.filter_by(ativo=True).order_by(
            ContaBancaria.principal.desc(),
            ContaBancaria.nome.asc()
        ).all()
        
        # Calcular saldo total
        saldo_total = sum(float(conta.saldo_atual) for conta in contas)
        saldo_com_limite = sum(conta.saldo_com_limite for conta in contas)
        
        # Estatísticas
        contas_ativas = len([c for c in contas if c.ativa])
        contas_inativas = len([c for c in contas if not c.ativa])
        
        return render_template('financeiro/contas_bancarias/listar.html',
                             contas=contas,
                             saldo_total=saldo_total,
                             saldo_com_limite=saldo_com_limite,
                             contas_ativas=contas_ativas,
                             contas_inativas=contas_inativas)
    
    except Exception as e:
        flash(f'Erro ao listar contas bancárias: {str(e)}', 'danger')
        return redirect(url_for('financeiro.dashboard'))


@bp_financeiro.route('/contas-bancarias/nova')
def nova_conta_bancaria():
    """Formulário para nova conta bancária."""
    return render_template('financeiro/contas_bancarias/form.html')


@bp_financeiro.route('/contas-bancarias/criar', methods=['POST'])
def criar_conta_bancaria():
    """Cria nova conta bancária."""
    try:
        # Dados do formulário
        nome = request.form.get('nome')
        tipo = request.form.get('tipo')
        banco = request.form.get('banco')
        agencia = request.form.get('agencia')
        numero_conta = request.form.get('numero_conta')
        saldo_inicial_str = request.form.get('saldo_inicial')
        limite_credito_str = request.form.get('limite_credito')
        principal = bool(request.form.get('principal'))
        ativa = bool(request.form.get('ativa', True))
        observacoes = request.form.get('observacoes')
        
        # Validações
        if not nome:
            flash('Nome da conta é obrigatório', 'danger')
            return redirect(url_for('financeiro.nova_conta_bancaria'))
        
        if not tipo:
            flash('Tipo da conta é obrigatório', 'danger')
            return redirect(url_for('financeiro.nova_conta_bancaria'))
        
        # Conversões
        saldo_inicial = converter_valor_monetario(saldo_inicial_str) if saldo_inicial_str else Decimal('0.00')
        limite_credito = converter_valor_monetario(limite_credito_str) if limite_credito_str else Decimal('0.00')
        
        # Se esta conta for principal, remover flag de outras
        if principal:
            ContaBancaria.query.filter_by(principal=True).update({'principal': False})
        
        # Criar conta
        conta = ContaBancaria(
            nome=nome,
            tipo=tipo,
            banco=banco,
            agencia=agencia,
            numero_conta=numero_conta,
            saldo_inicial=saldo_inicial,
            saldo_atual=saldo_inicial,  # Saldo atual começa igual ao inicial
            limite_credito=limite_credito,
            principal=principal,
            ativa=ativa,
            observacoes=observacoes
        )
        
        db.session.add(conta)
        db.session.commit()
        
        flash(f'Conta bancária "{nome}" criada com sucesso!', 'success')
        return redirect(url_for('financeiro.listar_contas_bancarias'))
    
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao criar conta bancária: {str(e)}', 'danger')
        return redirect(url_for('financeiro.nova_conta_bancaria'))


@bp_financeiro.route('/contas-bancarias/<int:id>/editar')
def editar_conta_bancaria(id):
    """Formulário para editar conta bancária."""
    try:
        conta = ContaBancaria.query.get_or_404(id)
        return render_template('financeiro/contas_bancarias/form.html', conta=conta)
    
    except Exception as e:
        flash(f'Erro ao carregar conta bancária: {str(e)}', 'danger')
        return redirect(url_for('financeiro.listar_contas_bancarias'))


@bp_financeiro.route('/contas-bancarias/<int:id>/atualizar', methods=['POST'])
def atualizar_conta_bancaria(id):
    """Atualiza conta bancária."""
    try:
        conta = ContaBancaria.query.get_or_404(id)
        
        # Dados do formulário
        conta.nome = request.form.get('nome')
        conta.tipo = request.form.get('tipo')
        conta.banco = request.form.get('banco')
        conta.agencia = request.form.get('agencia')
        conta.numero_conta = request.form.get('numero_conta')
        
        # Saldo inicial - só atualizar se mudou
        saldo_inicial_str = request.form.get('saldo_inicial')
        novo_saldo_inicial = converter_valor_monetario(saldo_inicial_str) if saldo_inicial_str else Decimal('0.00')
        
        # Se saldo inicial mudou, ajustar saldo atual proporcionalmente
        if novo_saldo_inicial != conta.saldo_inicial:
            diferenca = novo_saldo_inicial - conta.saldo_inicial
            conta.saldo_atual += diferenca
            conta.saldo_inicial = novo_saldo_inicial
        
        limite_credito_str = request.form.get('limite_credito')
        conta.limite_credito = converter_valor_monetario(limite_credito_str) if limite_credito_str else Decimal('0.00')
        
        principal = bool(request.form.get('principal'))
        conta.ativa = bool(request.form.get('ativa', True))
        conta.observacoes = request.form.get('observacoes')
        
        # Se esta conta for principal, remover flag de outras
        if principal and not conta.principal:
            ContaBancaria.query.filter(ContaBancaria.id != id, ContaBancaria.principal == True).update({'principal': False})
        
        conta.principal = principal
        
        # Validações
        if not conta.nome:
            flash('Nome da conta é obrigatório', 'danger')
            return redirect(url_for('financeiro.editar_conta_bancaria', id=id))
        
        if not conta.tipo:
            flash('Tipo da conta é obrigatório', 'danger')
            return redirect(url_for('financeiro.editar_conta_bancaria', id=id))
        
        db.session.commit()
        
        flash(f'Conta bancária "{conta.nome}" atualizada com sucesso!', 'success')
        return redirect(url_for('financeiro.listar_contas_bancarias'))
    
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar conta bancária: {str(e)}', 'danger')
        return redirect(url_for('financeiro.editar_conta_bancaria', id=id))


@bp_financeiro.route('/contas-bancarias/<int:id>/excluir', methods=['POST'])
def excluir_conta_bancaria(id):
    """Exclui conta bancária (soft delete)."""
    try:
        conta = ContaBancaria.query.get_or_404(id)
        
        # Verificar se há lançamentos vinculados
        lancamentos_vinculados = LancamentoFinanceiro.query.filter_by(
            conta_bancaria_id=id,
            ativo=True
        ).count()
        
        if lancamentos_vinculados > 0:
            flash(f'Não é possível excluir. Existem {lancamentos_vinculados} lançamento(s) vinculado(s) a esta conta.', 'danger')
            return redirect(url_for('financeiro.listar_contas_bancarias'))
        
        # Soft delete
        conta.ativo = False
        conta.ativa = False
        conta.data_exclusao = datetime.utcnow()
        
        db.session.commit()
        
        flash(f'Conta bancária "{conta.nome}" excluída com sucesso!', 'success')
        return redirect(url_for('financeiro.listar_contas_bancarias'))
    
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir conta bancária: {str(e)}', 'danger')
        return redirect(url_for('financeiro.listar_contas_bancarias'))


@bp_financeiro.route('/contas-bancarias/dashboard')
def dashboard_contas():
    """Dashboard de contas bancárias com saldos e movimentações."""
    try:
        contas = ContaBancaria.query.filter_by(ativo=True, ativa=True).all()
        
        # Estatísticas gerais
        saldo_total = sum(float(conta.saldo_atual) for conta in contas)
        saldo_com_limite = sum(conta.saldo_com_limite for conta in contas)
        
        # Últimas movimentações por conta
        movimentacoes = {}
        for conta in contas:
            ultimas = LancamentoFinanceiro.query.filter_by(
                conta_bancaria_id=conta.id,
                ativo=True
            ).order_by(LancamentoFinanceiro.data_lancamento.desc()).limit(5).all()
            movimentacoes[conta.id] = ultimas
        
        # Projeção para próximos 30 dias
        fluxo_projetado = LancamentoFinanceiro.calcular_fluxo_caixa(30)
        
        return render_template('financeiro/contas_bancarias/dashboard.html',
                             contas=contas,
                             saldo_total=saldo_total,
                             saldo_com_limite=saldo_com_limite,
                             movimentacoes=movimentacoes,
                             fluxo_projetado=fluxo_projetado)
    
    except Exception as e:
        flash(f'Erro ao carregar dashboard: {str(e)}', 'danger')
        return redirect(url_for('financeiro.dashboard'))


@bp_financeiro.route('/contas-bancarias/transferencia')
def nova_transferencia():
    """Formulário para transferência entre contas."""
    try:
        contas = ContaBancaria.query.filter_by(ativo=True, ativa=True).all()
        return render_template('financeiro/contas_bancarias/transferencia.html', 
                             contas=contas,
                             data_hoje=date.today())
    
    except Exception as e:
        flash(f'Erro ao carregar formulário: {str(e)}', 'danger')
        return redirect(url_for('financeiro.listar_contas_bancarias'))


@bp_financeiro.route('/contas-bancarias/transferencia/executar', methods=['POST'])
def executar_transferencia():
    """Executa transferência entre contas bancárias."""
    try:
        # Dados do formulário
        conta_origem_id = request.form.get('conta_origem_id')
        conta_destino_id = request.form.get('conta_destino_id')
        valor_str = request.form.get('valor')
        data_transferencia_str = request.form.get('data_transferencia')
        descricao = request.form.get('descricao')
        
        # Validações
        if not conta_origem_id or not conta_destino_id:
            flash('Selecione as contas de origem e destino', 'danger')
            return redirect(url_for('financeiro.nova_transferencia'))
        
        if conta_origem_id == conta_destino_id:
            flash('As contas de origem e destino devem ser diferentes', 'danger')
            return redirect(url_for('financeiro.nova_transferencia'))
        
        if not valor_str:
            flash('Informe o valor da transferência', 'danger')
            return redirect(url_for('financeiro.nova_transferencia'))
        
        # Conversões
        valor = converter_valor_monetario(valor_str)
        
        if valor <= 0:
            flash('O valor deve ser maior que zero', 'danger')
            return redirect(url_for('financeiro.nova_transferencia'))
        
        data_transferencia = datetime.strptime(data_transferencia_str, '%Y-%m-%d').date() if data_transferencia_str else date.today()
        
        # Buscar contas
        conta_origem = ContaBancaria.query.get_or_404(int(conta_origem_id))
        conta_destino = ContaBancaria.query.get_or_404(int(conta_destino_id))
        
        # Verificar saldo suficiente
        if float(conta_origem.saldo_atual) < float(valor):
            flash(f'Saldo insuficiente na conta {conta_origem.nome}. Saldo disponível: {conta_origem.saldo_formatado}', 'danger')
            return redirect(url_for('financeiro.nova_transferencia'))
        
        # Criar lançamento de saída (débito)
        lancamento_saida = LancamentoFinanceiro(
            descricao=descricao or f'Transferência para {conta_destino.nome}',
            valor=valor,
            tipo='despesa',
            status='pago',
            categoria='Transferência Bancária',
            subcategoria='Débito',
            data_lancamento=data_transferencia,
            data_pagamento=data_transferencia,
            conta_bancaria_id=conta_origem.id,
            observacoes=f'Transferência para conta {conta_destino.nome} (ID: {conta_destino.id})'
        )
        
        # Criar lançamento de entrada (crédito)
        lancamento_entrada = LancamentoFinanceiro(
            descricao=descricao or f'Transferência de {conta_origem.nome}',
            valor=valor,
            tipo='receita',
            status='recebido',
            categoria='Transferência Bancária',
            subcategoria='Crédito',
            data_lancamento=data_transferencia,
            data_pagamento=data_transferencia,
            conta_bancaria_id=conta_destino.id,
            observacoes=f'Transferência da conta {conta_origem.nome} (ID: {conta_origem.id})'
        )
        
        # Atualizar saldos
        conta_origem.atualizar_saldo(valor, 'subtrair')
        conta_destino.atualizar_saldo(valor, 'adicionar')
        
        # Salvar lançamentos
        db.session.add(lancamento_saida)
        db.session.add(lancamento_entrada)
        db.session.commit()
        
        flash(f'Transferência de {formatar_valor_real(valor)} realizada com sucesso!', 'success')
        return redirect(url_for('financeiro.dashboard_contas'))
    
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao executar transferência: {str(e)}', 'danger')
        return redirect(url_for('financeiro.nova_transferencia'))


# ========================================
# CENTROS DE CUSTO
# ========================================

@bp_financeiro.route('/centros-custo')
def listar_centros_custo():
    """Lista todos os centros de custo."""
    try:
        centros = CentroCusto.query.filter_by(ativo=True).order_by(
            CentroCusto.codigo.asc()
        ).all()
        
        # Calcular total de despesas por centro (mês atual)
        hoje = date.today()
        mes_atual = hoje.month
        ano_atual = hoje.year
        
        despesas_por_centro = {}
        for centro in centros:
            despesas_mes = CentroCusto.get_despesas_mes(centro.id, mes_atual, ano_atual)
            despesas_por_centro[centro.id] = despesas_mes
            
            # Calcular percentual do orçamento se houver
            if centro.orcamento_mensal and centro.orcamento_mensal > 0:
                centro.percentual_usado = (despesas_mes / float(centro.orcamento_mensal)) * 100
            else:
                centro.percentual_usado = 0
        
        return render_template('financeiro/centros_custo/listar.html',
                             centros=centros,
                             despesas_por_centro=despesas_por_centro)
    
    except Exception as e:
        flash(f'Erro ao listar centros de custo: {str(e)}', 'danger')
        return redirect(url_for('financeiro.dashboard'))


@bp_financeiro.route('/centros-custo/novo')
def novo_centro_custo():
    """Formulário para novo centro de custo."""
    # Buscar centros para hierarquia
    centros_pais = CentroCusto.query.filter_by(ativo=True, centro_pai_id=None).all()
    return render_template('financeiro/centros_custo/form.html', centros_pais=centros_pais)


@bp_financeiro.route('/centros-custo/criar', methods=['POST'])
def criar_centro_custo():
    """Cria novo centro de custo."""
    try:
        # Dados do formulário
        codigo = request.form.get('codigo')
        nome = request.form.get('nome')
        descricao = request.form.get('descricao')
        tipo = request.form.get('tipo')
        responsavel = request.form.get('responsavel')
        orcamento_mensal_str = request.form.get('orcamento_mensal')
        centro_pai_id = request.form.get('centro_pai_id')
        
        # Validações
        if not codigo:
            flash('Código é obrigatório', 'danger')
            return redirect(url_for('financeiro.novo_centro_custo'))
        
        if not nome:
            flash('Nome é obrigatório', 'danger')
            return redirect(url_for('financeiro.novo_centro_custo'))
        
        # Verificar se código já existe
        codigo_existente = CentroCusto.query.filter_by(codigo=codigo, ativo=True).first()
        if codigo_existente:
            flash(f'Já existe um centro de custo com o código "{codigo}"', 'danger')
            return redirect(url_for('financeiro.novo_centro_custo'))
        
        # Conversões
        orcamento_mensal = converter_valor_monetario(orcamento_mensal_str) if orcamento_mensal_str else None
        
        # Criar centro de custo
        centro = CentroCusto(
            codigo=codigo,
            nome=nome,
            descricao=descricao,
            tipo=tipo,
            responsavel=responsavel,
            orcamento_mensal=orcamento_mensal,
            centro_pai_id=int(centro_pai_id) if centro_pai_id else None
        )
        
        db.session.add(centro)
        db.session.commit()
        
        flash(f'Centro de custo "{codigo} - {nome}" criado com sucesso!', 'success')
        return redirect(url_for('financeiro.listar_centros_custo'))
    
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao criar centro de custo: {str(e)}', 'danger')
        return redirect(url_for('financeiro.novo_centro_custo'))


@bp_financeiro.route('/centros-custo/<int:id>/editar')
def editar_centro_custo(id):
    """Formulário para editar centro de custo."""
    try:
        centro = CentroCusto.query.get_or_404(id)
        centros_pais = CentroCusto.query.filter(
            CentroCusto.ativo == True,
            CentroCusto.id != id,
            CentroCusto.centro_pai_id == None
        ).all()
        return render_template('financeiro/centros_custo/form.html', 
                             centro=centro,
                             centros_pais=centros_pais)
    
    except Exception as e:
        flash(f'Erro ao carregar centro de custo: {str(e)}', 'danger')
        return redirect(url_for('financeiro.listar_centros_custo'))


@bp_financeiro.route('/centros-custo/<int:id>/atualizar', methods=['POST'])
def atualizar_centro_custo(id):
    """Atualiza centro de custo."""
    try:
        centro = CentroCusto.query.get_or_404(id)
        
        # Dados do formulário
        codigo = request.form.get('codigo')
        centro.nome = request.form.get('nome')
        centro.descricao = request.form.get('descricao')
        centro.tipo = request.form.get('tipo')
        centro.responsavel = request.form.get('responsavel')
        orcamento_mensal_str = request.form.get('orcamento_mensal')
        centro_pai_id = request.form.get('centro_pai_id')
        
        # Validações
        if not codigo:
            flash('Código é obrigatório', 'danger')
            return redirect(url_for('financeiro.editar_centro_custo', id=id))
        
        if not centro.nome:
            flash('Nome é obrigatório', 'danger')
            return redirect(url_for('financeiro.editar_centro_custo', id=id))
        
        # Verificar se código já existe em outro centro
        if codigo != centro.codigo:
            codigo_existente = CentroCusto.query.filter(
                CentroCusto.codigo == codigo,
                CentroCusto.ativo == True,
                CentroCusto.id != id
            ).first()
            if codigo_existente:
                flash(f'Já existe outro centro de custo com o código "{codigo}"', 'danger')
                return redirect(url_for('financeiro.editar_centro_custo', id=id))
            centro.codigo = codigo
        
        # Conversões
        centro.orcamento_mensal = converter_valor_monetario(orcamento_mensal_str) if orcamento_mensal_str else None
        centro.centro_pai_id = int(centro_pai_id) if centro_pai_id else None
        
        db.session.commit()
        
        flash(f'Centro de custo "{centro.codigo} - {centro.nome}" atualizado com sucesso!', 'success')
        return redirect(url_for('financeiro.listar_centros_custo'))
    
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar centro de custo: {str(e)}', 'danger')
        return redirect(url_for('financeiro.editar_centro_custo', id=id))


@bp_financeiro.route('/centros-custo/<int:id>/excluir', methods=['POST'])
def excluir_centro_custo(id):
    """Exclui centro de custo (soft delete)."""
    try:
        centro = CentroCusto.query.get_or_404(id)
        
        # Verificar se há lançamentos vinculados
        lancamentos_vinculados = LancamentoFinanceiro.query.filter_by(
            centro_custo_id=id,
            ativo=True
        ).count()
        
        if lancamentos_vinculados > 0:
            flash(f'Não é possível excluir. Existem {lancamentos_vinculados} lançamento(s) vinculado(s) a este centro de custo.', 'danger')
            return redirect(url_for('financeiro.listar_centros_custo'))
        
        # Verificar se há centros filhos
        centros_filhos = CentroCusto.query.filter_by(
            centro_pai_id=id,
            ativo=True
        ).count()
        
        if centros_filhos > 0:
            flash(f'Não é possível excluir. Existem {centros_filhos} centro(s) de custo subordinado(s).', 'danger')
            return redirect(url_for('financeiro.listar_centros_custo'))
        
        # Soft delete
        centro.ativo = False
        centro.data_exclusao = datetime.utcnow()
        
        db.session.commit()
        
        flash(f'Centro de custo "{centro.codigo} - {centro.nome}" excluído com sucesso!', 'success')
        return redirect(url_for('financeiro.listar_centros_custo'))
    
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir centro de custo: {str(e)}', 'danger')
        return redirect(url_for('financeiro.listar_centros_custo'))


@bp_financeiro.route('/centros-custo/<int:id>/relatorio')
def relatorio_centro_custo(id):
    """Relatório detalhado de um centro de custo."""
    try:
        centro = CentroCusto.query.get_or_404(id)
        
        # Filtros de período
        mes = request.args.get('mes', type=int) or date.today().month
        ano = request.args.get('ano', type=int) or date.today().year
        
        # Buscar todos os lançamentos do centro no período
        lancamentos = LancamentoFinanceiro.query.filter(
            LancamentoFinanceiro.centro_custo_id == id,
            LancamentoFinanceiro.ativo == True,
            db.extract('month', LancamentoFinanceiro.data_lancamento) == mes,
            db.extract('year', LancamentoFinanceiro.data_lancamento) == ano
        ).order_by(LancamentoFinanceiro.data_lancamento.desc()).all()
        
        # Calcular totais
        total_receitas = sum(float(l.valor) for l in lancamentos if l.tipo in ['receita', 'conta_receber'])
        total_despesas = sum(float(l.valor) for l in lancamentos if l.tipo in ['despesa', 'conta_pagar'])
        saldo = total_receitas - total_despesas
        
        # Calcular por categoria
        despesas_por_categoria = {}
        for lanc in lancamentos:
            if lanc.tipo in ['despesa', 'conta_pagar'] and lanc.categoria:
                if lanc.categoria not in despesas_por_categoria:
                    despesas_por_categoria[lanc.categoria] = 0
                despesas_por_categoria[lanc.categoria] += float(lanc.valor)
        
        # Ordenar por valor
        despesas_por_categoria = dict(sorted(despesas_por_categoria.items(), 
                                            key=lambda x: x[1], reverse=True))
        
        # Percentual do orçamento
        percentual_orcamento = 0
        if centro.orcamento_mensal and centro.orcamento_mensal > 0:
            percentual_orcamento = (total_despesas / float(centro.orcamento_mensal)) * 100
        
        return render_template('financeiro/centros_custo/relatorio.html',
                             centro=centro,
                             lancamentos=lancamentos,
                             total_receitas=total_receitas,
                             total_despesas=total_despesas,
                             saldo=saldo,
                             despesas_por_categoria=despesas_por_categoria,
                             percentual_orcamento=percentual_orcamento,
                             mes=mes,
                             ano=ano)
    
    except Exception as e:
        flash(f'Erro ao gerar relatório: {str(e)}', 'danger')
        return redirect(url_for('financeiro.listar_centros_custo'))


# Importar formatar_valor_real
from app.financeiro.financeiro_utils import formatar_valor_real


# =============================================================================
# CONCILIAÇÃO BANCÁRIA
# =============================================================================

@bp_financeiro.route('/conciliacao-bancaria')
def conciliacao_bancaria():
    """Página principal de conciliação bancária."""
    try:
        # Buscar contas bancárias ativas
        contas = ContaBancaria.query.filter_by(ativo=True).all()
        
        # Conta selecionada (via query string)
        conta_id = request.args.get('conta_id', type=int)
        conta_selecionada = None
        extratos_pendentes = []
        lancamentos_pendentes = []
        
        if conta_id:
            conta_selecionada = ContaBancaria.query.get(conta_id)
            if conta_selecionada:
                # Buscar extratos pendentes desta conta
                from app.financeiro.financeiro_model import ExtratoBancario
                extratos_pendentes = ExtratoBancario.get_pendentes(conta_id).all()
                
                # Buscar lançamentos não conciliados desta conta
                lancamentos_pendentes = LancamentoFinanceiro.query.filter_by(
                    conta_bancaria_id=conta_id,
                    ativo=True
                ).filter(
                    ~LancamentoFinanceiro.id.in_(
                        db.session.query(ExtratoBancario.lancamento_id).filter(
                            ExtratoBancario.lancamento_id.isnot(None)
                        )
                    )
                ).order_by(LancamentoFinanceiro.data_vencimento.desc()).limit(50).all()
        
        return render_template('financeiro/conciliacao_bancaria/conciliacao.html',
                             contas=contas,
                             conta_selecionada=conta_selecionada,
                             extratos_pendentes=extratos_pendentes,
                             lancamentos_pendentes=lancamentos_pendentes)
    
    except Exception as e:
        flash(f'Erro ao carregar conciliação: {str(e)}', 'danger')
        return redirect(url_for('financeiro.dashboard'))


@bp_financeiro.route('/conciliacao-bancaria/upload', methods=['GET', 'POST'])
def upload_extrato():
    """Upload de arquivo de extrato bancário."""
    if request.method == 'POST':
        try:
            # Validar arquivo
            if 'arquivo' not in request.files:
                flash('Nenhum arquivo selecionado.', 'warning')
                return redirect(url_for('financeiro.upload_extrato'))
            
            arquivo = request.files['arquivo']
            if arquivo.filename == '':
                flash('Nenhum arquivo selecionado.', 'warning')
                return redirect(url_for('financeiro.upload_extrato'))
            
            # Validar conta bancária
            conta_id = request.form.get('conta_bancaria_id', type=int)
            if not conta_id:
                flash('Selecione uma conta bancária.', 'warning')
                return redirect(url_for('financeiro.upload_extrato'))
            
            conta = ContaBancaria.query.get(conta_id)
            if not conta:
                flash('Conta bancária não encontrada.', 'danger')
                return redirect(url_for('financeiro.upload_extrato'))
            
            # Processar arquivo CSV
            import csv
            import io
            from app.financeiro.financeiro_model import ExtratoBancario
            
            # Ler arquivo
            stream = io.StringIO(arquivo.stream.read().decode("UTF-8"), newline=None)
            csv_reader = csv.DictReader(stream)
            
            # Formato esperado das colunas (pode variar por banco)
            # data,descricao,documento,valor,tipo
            
            contador = 0
            for row in csv_reader:
                try:
                    # Parsear data (formato DD/MM/YYYY)
                    data_str = row.get('data', '').strip()
                    if '/' in data_str:
                        dia, mes, ano = data_str.split('/')
                        data_movimento = date(int(ano), int(mes), int(dia))
                    else:
                        continue  # Pular linha sem data válida
                    
                    # Parsear descrição
                    descricao = row.get('descricao', '').strip()
                    if not descricao:
                        continue
                    
                    # Parsear valor
                    valor_str = row.get('valor', '0').strip()
                    valor = converter_valor_monetario(valor_str)
                    
                    # Tipo de movimento
                    tipo = row.get('tipo', 'debito').strip().lower()
                    if tipo not in ['debito', 'credito']:
                        tipo = 'debito' if valor < 0 else 'credito'
                    
                    # Documento (opcional)
                    documento = row.get('documento', '').strip()
                    
                    # Criar extrato
                    extrato = ExtratoBancario(
                        conta_bancaria_id=conta_id,
                        data_movimento=data_movimento,
                        descricao=descricao,
                        documento=documento,
                        valor=abs(valor),
                        tipo_movimento=tipo,
                        arquivo_origem=arquivo.filename
                    )
                    db.session.add(extrato)
                    contador += 1
                
                except Exception as e:
                    continue  # Pular linhas com erro
            
            db.session.commit()
            flash(f'Extrato importado com sucesso! {contador} lançamentos adicionados.', 'success')
            return redirect(url_for('financeiro.conciliacao_bancaria', conta_id=conta_id))
        
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao processar arquivo: {str(e)}', 'danger')
            return redirect(url_for('financeiro.upload_extrato'))
    
    # GET - Exibir formulário
    contas = ContaBancaria.query.filter_by(ativo=True).all()
    return render_template('financeiro/conciliacao_bancaria/upload.html', contas=contas)


@bp_financeiro.route('/conciliacao-bancaria/conciliar/<int:extrato_id>/<int:lancamento_id>', methods=['POST'])
def conciliar_manual(extrato_id, lancamento_id):
    """Conciliar manualmente um extrato com um lançamento."""
    try:
        from app.financeiro.financeiro_model import ExtratoBancario
        
        extrato = ExtratoBancario.query.get_or_404(extrato_id)
        lancamento = LancamentoFinanceiro.query.get_or_404(lancamento_id)
        
        # Verificar se já não está conciliado
        if extrato.conciliado:
            flash('Este extrato já está conciliado.', 'warning')
            return redirect(url_for('financeiro.conciliacao_bancaria', conta_id=extrato.conta_bancaria_id))
        
        # Realizar conciliação
        extrato.conciliar_com_lancamento(lancamento.id)
        
        flash('Conciliação realizada com sucesso!', 'success')
        return redirect(url_for('financeiro.conciliacao_bancaria', conta_id=extrato.conta_bancaria_id))
    
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao conciliar: {str(e)}', 'danger')
        return redirect(url_for('financeiro.conciliacao_bancaria'))


@bp_financeiro.route('/conciliacao-bancaria/desconciliar/<int:extrato_id>', methods=['POST'])
def desconciliar(extrato_id):
    """Desfazer conciliação de um extrato."""
    try:
        from app.financeiro.financeiro_model import ExtratoBancario
        
        extrato = ExtratoBancario.query.get_or_404(extrato_id)
        conta_id = extrato.conta_bancaria_id
        
        # Desfazer conciliação
        extrato.desconciliar()
        
        flash('Conciliação desfeita com sucesso!', 'success')
        return redirect(url_for('financeiro.conciliacao_bancaria', conta_id=conta_id))
    
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao desconciliar: {str(e)}', 'danger')
        return redirect(url_for('financeiro.conciliacao_bancaria'))


@bp_financeiro.route('/conciliacao-bancaria/historico')
def historico_conciliacao():
    """Histórico de conciliações realizadas."""
    try:
        from app.financeiro.financeiro_model import ExtratoBancario
        
        # Filtros
        conta_id = request.args.get('conta_id', type=int)
        
        # Buscar extratos conciliados
        query = ExtratoBancario.get_conciliados(conta_id)
        extratos = query.limit(100).all()
        
        # Buscar contas para filtro
        contas = ContaBancaria.query.filter_by(ativo=True).all()
        
        return render_template('financeiro/conciliacao_bancaria/historico.html',
                             extratos=extratos,
                             contas=contas,
                             conta_id=conta_id)
    
    except Exception as e:
        flash(f'Erro ao carregar histórico: {str(e)}', 'danger')
        return redirect(url_for('financeiro.dashboard'))


# ==================== ROTAS DE CUSTOS FIXOS ====================

@bp_financeiro.route('/custos-fixos')
def listar_custos_fixos():
    """Lista todos os custos fixos."""
    try:
        from app.financeiro.financeiro_model import CustoFixo
        
        # Filtros
        categoria = request.args.get('categoria')
        status = request.args.get('status', 'ativo')
        
        # Query base
        query = CustoFixo.query
        
        if status == 'ativo':
            query = query.filter_by(ativo=True)
        elif status == 'inativo':
            query = query.filter_by(ativo=False)
        
        if categoria:
            query = query.filter_by(categoria=categoria)
        
        custos = query.order_by(CustoFixo.dia_vencimento).all()
        
        # Calcular total mensal
        total_mensal = CustoFixo.get_total_mensal()
        
        # Buscar categorias únicas
        categorias = db.session.query(CustoFixo.categoria).distinct().all()
        categorias = [c[0] for c in categorias if c[0]]
        
        return render_template('financeiro/custos_fixos/listar.html',
                             custos=custos,
                             total_mensal=total_mensal,
                             categorias=categorias,
                             categoria_filtro=categoria,
                             status_filtro=status)
    
    except Exception as e:
        flash(f'Erro ao listar custos fixos: {str(e)}', 'danger')
        return redirect(url_for('financeiro.dashboard'))


@bp_financeiro.route('/custos-fixos/novo', methods=['GET', 'POST'])
def novo_custo_fixo():
    """Formulário para criar novo custo fixo."""
    try:
        from app.financeiro.financeiro_model import CustoFixo
        
        if request.method == 'POST':
            try:
                # Criar custo fixo
                custo = CustoFixo(
                    nome=request.form['nome'],
                    descricao=request.form.get('descricao', ''),
                    valor_mensal=converter_valor_monetario(request.form['valor_mensal']),
                    categoria=request.form['categoria'],
                    tipo=request.form.get('tipo', 'DESPESA'),
                    dia_vencimento=int(request.form['dia_vencimento']),
                    gerar_automaticamente=bool(request.form.get('gerar_automaticamente')),
                    data_inicio=datetime.strptime(request.form['data_inicio'], '%Y-%m-%d').date(),
                    data_fim=datetime.strptime(request.form['data_fim'], '%Y-%m-%d').date() if request.form.get('data_fim') else None,
                    conta_bancaria_id=int(request.form['conta_bancaria_id']) if request.form.get('conta_bancaria_id') else None,
                    centro_custo_id=int(request.form['centro_custo_id']) if request.form.get('centro_custo_id') else None,
                    ativo=True
                )
                
                db.session.add(custo)
                db.session.commit()
                
                logger.info(f'✅ Custo fixo criado: {custo.nome} (ID: {custo.id})')
                flash(f'Custo fixo "{custo.nome}" criado com sucesso!', 'success')
                
                # Redirecionar diretamente para a lista
                redirect_url = url_for('financeiro.listar_custos_fixos')
                logger.info(f'🔄 Redirecionando para: {redirect_url}')
                return redirect(redirect_url)
            
            except Exception as post_error:
                db.session.rollback()
                logger.error(f'❌ Erro no POST do custo fixo: {str(post_error)}')
                logger.exception(post_error)
                flash(f'Erro ao salvar custo fixo: {str(post_error)}', 'danger')
                # Em caso de erro, recarregar o formulário
                contas = ContaBancaria.query.filter_by(ativo=True).all()
                centros = CentroCusto.query.filter_by(ativo=True).all()
                categorias_padrao = [
                    'Aluguel', 'Salários', 'Encargos', 'Energia', 'Água',
                    'Internet', 'Telefone', 'Software', 'Manutenção', 'Seguros',
                    'Impostos', 'Contabilidade', 'Marketing', 'Outros'
                ]
                return render_template('financeiro/custos_fixos/form.html',
                                     custo=None,
                                     contas=contas,
                                     centros=centros,
                                     categorias=categorias_padrao,
                                     data_hoje=date.today())
        
        # GET - exibir formulário
        contas = ContaBancaria.query.filter_by(ativo=True).all()
        centros = CentroCusto.query.filter_by(ativo=True).all()
        
        # Categorias padrão
        categorias_padrao = [
            'Aluguel', 'Salários', 'Encargos', 'Energia', 'Água',
            'Internet', 'Telefone', 'Software', 'Manutenção', 'Seguros',
            'Impostos', 'Contabilidade', 'Marketing', 'Outros'
        ]
        
        return render_template('financeiro/custos_fixos/form.html',
                             custo=None,
                             contas=contas,
                             centros=centros,
                             categorias=categorias_padrao,
                             data_hoje=date.today())
    
    except Exception as e:
        db.session.rollback()
        logger.error(f'❌ Erro ao criar custo fixo: {str(e)}')
        logger.exception(e)  # Log completo do stack trace
        flash(f'Erro ao criar custo fixo: {str(e)}', 'danger')
        return redirect(url_for('financeiro.listar_custos_fixos'))


@bp_financeiro.route('/custos-fixos/<int:id>/editar', methods=['GET', 'POST'])
def editar_custo_fixo(id):
    """Editar custo fixo existente."""
    try:
        from app.financeiro.financeiro_model import CustoFixo
        
        custo = CustoFixo.query.get_or_404(id)
        
        if request.method == 'POST':
            # Atualizar custo
            custo.nome = request.form['nome']
            custo.descricao = request.form.get('descricao', '')
            custo.valor_mensal = converter_valor_monetario(request.form['valor_mensal'])
            custo.categoria = request.form['categoria']
            custo.tipo = request.form.get('tipo', 'DESPESA')
            custo.dia_vencimento = int(request.form['dia_vencimento'])
            custo.gerar_automaticamente = bool(request.form.get('gerar_automaticamente'))
            custo.data_inicio = datetime.strptime(request.form['data_inicio'], '%Y-%m-%d').date()
            custo.data_fim = datetime.strptime(request.form['data_fim'], '%Y-%m-%d').date() if request.form.get('data_fim') else None
            custo.conta_bancaria_id = int(request.form['conta_bancaria_id']) if request.form.get('conta_bancaria_id') else None
            custo.centro_custo_id = int(request.form['centro_custo_id']) if request.form.get('centro_custo_id') else None
            
            db.session.commit()
            
            flash(f'Custo fixo "{custo.nome}" atualizado com sucesso!', 'success')
            return redirect(url_for('financeiro.listar_custos_fixos'))
        
        # GET - exibir formulário
        contas = ContaBancaria.query.filter_by(ativo=True).all()
        centros = CentroCusto.query.filter_by(ativo=True).all()
        
        categorias_padrao = [
            'Aluguel', 'Salários', 'Encargos', 'Energia', 'Água',
            'Internet', 'Telefone', 'Software', 'Manutenção', 'Seguros',
            'Impostos', 'Contabilidade', 'Marketing', 'Outros'
        ]
        
        return render_template('financeiro/custos_fixos/form.html',
                             custo=custo,
                             contas=contas,
                             centros=centros,
                             categorias=categorias_padrao,
                             data_hoje=date.today())
    
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao editar custo fixo: {str(e)}', 'danger')
        return redirect(url_for('financeiro.listar_custos_fixos'))


@bp_financeiro.route('/custos-fixos/<int:id>/excluir', methods=['POST'])
def excluir_custo_fixo(id):
    """Excluir (desativar) custo fixo."""
    try:
        from app.financeiro.financeiro_model import CustoFixo
        
        custo = CustoFixo.query.get_or_404(id)
        custo.ativo = False
        
        db.session.commit()
        
        flash(f'Custo fixo "{custo.nome}" desativado com sucesso!', 'success')
    
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir custo fixo: {str(e)}', 'danger')
    
    return redirect(url_for('financeiro.listar_custos_fixos'))


@bp_financeiro.route('/custos-fixos/dashboard')
def dashboard_custos_fixos():
    """Dashboard de custos fixos."""
    try:
        from app.financeiro.financeiro_model import CustoFixo
        
        custos = CustoFixo.get_custos_ativos()
        total_mensal = CustoFixo.get_total_mensal()
        
        # Próximos vencimentos (próximos 30 dias)
        proximos_vencimentos = []
        for custo in custos:
            vencimento = custo.proximo_vencimento
            dias_restantes = (vencimento - date.today()).days
            if dias_restantes <= 30:
                proximos_vencimentos.append({
                    'custo': custo,
                    'vencimento': vencimento,
                    'dias_restantes': dias_restantes
                })
        
        proximos_vencimentos.sort(key=lambda x: x['vencimento'])
        
        # Totais por categoria
        totais_categoria = {}
        for custo in custos:
            if custo.categoria not in totais_categoria:
                totais_categoria[custo.categoria] = Decimal('0.00')
            totais_categoria[custo.categoria] += custo.valor_mensal
        
        return render_template('financeiro/custos_fixos/dashboard.html',
                             custos=custos,
                             total_mensal=total_mensal,
                             proximos_vencimentos=proximos_vencimentos,
                             totais_categoria=totais_categoria)
    
    except Exception as e:
        flash(f'Erro ao carregar dashboard: {str(e)}', 'danger')
        return redirect(url_for('financeiro.dashboard'))


@bp_financeiro.route('/custos-fixos/gerar-lancamentos', methods=['POST'])
def gerar_lancamentos_mes():
    """Gera lançamentos automáticos para o mês atual."""
    try:
        from app.financeiro.financeiro_model import CustoFixo
        
        lancamentos = CustoFixo.gerar_lancamentos_automaticos()
        
        if lancamentos:
            flash(f'{len(lancamentos)} lançamento(s) gerado(s) com sucesso!', 'success')
        else:
            flash('Nenhum lançamento novo para gerar.', 'info')
    
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao gerar lançamentos: {str(e)}', 'danger')
    
    return redirect(url_for('financeiro.dashboard_custos_fixos'))


# ==================== ROTAS DE FLUXO DE CAIXA ====================

@bp_financeiro.route('/fluxo-caixa')
def fluxo_caixa():
    """Dashboard de Fluxo de Caixa Projetado."""
    try:
        from datetime import timedelta
        from app.financeiro.financeiro_model import CustoFixo
        
        # Filtros
        conta_id = request.args.get('conta_id', type=int)
        periodo = request.args.get('periodo', '30')  # 30, 60, 90 dias
        
        # Data inicial e final
        data_hoje = date.today()
        if periodo == '30':
            data_fim = data_hoje + timedelta(days=30)
        elif periodo == '60':
            data_fim = data_hoje + timedelta(days=60)
        elif periodo == '90':
            data_fim = data_hoje + timedelta(days=90)
        else:
            data_fim = data_hoje + timedelta(days=30)
        
        # Saldo inicial (soma de todas as contas ou conta específica)
        query_saldo = ContaBancaria.query.filter_by(ativo=True)
        if conta_id:
            query_saldo = query_saldo.filter_by(id=conta_id)
        
        contas = query_saldo.all()
        saldo_inicial = sum(conta.saldo_atual for conta in contas)
        
        # Buscar lançamentos futuros
        query_lancamentos = LancamentoFinanceiro.query.filter(
            LancamentoFinanceiro.data_vencimento >= data_hoje,
            LancamentoFinanceiro.data_vencimento <= data_fim,
            LancamentoFinanceiro.ativo == True
        )
        
        if conta_id:
            query_lancamentos = query_lancamentos.filter_by(conta_bancaria_id=conta_id)
        
        lancamentos = query_lancamentos.order_by(LancamentoFinanceiro.data_vencimento).all()
        
        # Calcular projeção dia a dia
        projecao = []
        saldo_acumulado = saldo_inicial
        data_atual = data_hoje
        
        while data_atual <= data_fim:
            # Lançamentos do dia
            lancamentos_dia = [l for l in lancamentos if l.data_vencimento == data_atual]
            
            receitas_dia = sum(l.valor for l in lancamentos_dia if l.tipo == 'RECEITA')
            despesas_dia = sum(l.valor for l in lancamentos_dia if l.tipo == 'DESPESA')
            saldo_dia = receitas_dia - despesas_dia
            saldo_acumulado += saldo_dia
            
            projecao.append({
                'data': data_atual,
                'receitas': receitas_dia,
                'despesas': despesas_dia,
                'saldo_dia': saldo_dia,
                'saldo_acumulado': saldo_acumulado,
                'lancamentos': lancamentos_dia
            })
            
            data_atual += timedelta(days=1)
        
        # Identificar períodos de saldo negativo
        alertas = [p for p in projecao if p['saldo_acumulado'] < 0]
        
        # Totais do período
        total_receitas = sum(p['receitas'] for p in projecao)
        total_despesas = sum(p['despesas'] for p in projecao)
        saldo_final = projecao[-1]['saldo_acumulado'] if projecao else saldo_inicial
        
        # Buscar todas as contas para filtro
        todas_contas = ContaBancaria.query.filter_by(ativo=True).all()
        
        return render_template('financeiro/fluxo_caixa/dashboard.html',
                             projecao=projecao,
                             saldo_inicial=saldo_inicial,
                             saldo_final=saldo_final,
                             total_receitas=total_receitas,
                             total_despesas=total_despesas,
                             alertas=alertas,
                             contas=todas_contas,
                             conta_id=conta_id,
                             periodo=periodo,
                             data_hoje=data_hoje,
                             data_fim=data_fim)
    
    except Exception as e:
        flash(f'Erro ao carregar fluxo de caixa: {str(e)}', 'danger')
        return redirect(url_for('financeiro.dashboard'))


@bp_financeiro.route('/fluxo-caixa/exportar-excel')
def exportar_fluxo_excel():
    """Exportar fluxo de caixa para Excel."""
    try:
        from datetime import timedelta
        from io import BytesIO
        from flask import send_file
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Filtros
        conta_id = request.args.get('conta_id', type=int)
        periodo = request.args.get('periodo', '30')
        
        # Data inicial e final
        data_hoje = date.today()
        if periodo == '30':
            data_fim = data_hoje + timedelta(days=30)
        elif periodo == '60':
            data_fim = data_hoje + timedelta(days=60)
        else:
            data_fim = data_hoje + timedelta(days=90)
        
        # Saldo inicial
        query_saldo = ContaBancaria.query.filter_by(ativo=True)
        if conta_id:
            query_saldo = query_saldo.filter_by(id=conta_id)
        
        contas = query_saldo.all()
        saldo_inicial = sum(conta.saldo_atual for conta in contas)
        
        # Buscar lançamentos
        query_lancamentos = LancamentoFinanceiro.query.filter(
            LancamentoFinanceiro.data_vencimento >= data_hoje,
            LancamentoFinanceiro.data_vencimento <= data_fim,
            LancamentoFinanceiro.ativo == True
        )
        
        if conta_id:
            query_lancamentos = query_lancamentos.filter_by(conta_bancaria_id=conta_id)
        
        lancamentos = query_lancamentos.order_by(LancamentoFinanceiro.data_vencimento).all()
        
        # Criar Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fluxo de Caixa"
        
        # Cabeçalho
        ws.append(['FLUXO DE CAIXA PROJETADO'])
        ws.append([f'Período: {data_hoje.strftime("%d/%m/%Y")} a {data_fim.strftime("%d/%m/%Y")}'])
        ws.append([f'Saldo Inicial: R$ {saldo_inicial:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')])
        ws.append([])
        
        # Títulos das colunas
        ws.append(['Data', 'Receitas', 'Despesas', 'Saldo do Dia', 'Saldo Acumulado'])
        
        # Estilizar cabeçalho
        for cell in ws[5]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Dados
        saldo_acumulado = saldo_inicial
        data_atual = data_hoje
        
        while data_atual <= data_fim:
            lancamentos_dia = [l for l in lancamentos if l.data_vencimento == data_atual]
            
            receitas_dia = sum(l.valor for l in lancamentos_dia if l.tipo == 'RECEITA')
            despesas_dia = sum(l.valor for l in lancamentos_dia if l.tipo == 'DESPESA')
            saldo_dia = receitas_dia - despesas_dia
            saldo_acumulado += saldo_dia
            
            ws.append([
                data_atual.strftime('%d/%m/%Y'),
                float(receitas_dia),
                float(despesas_dia),
                float(saldo_dia),
                float(saldo_acumulado)
            ])
            
            # Colorir linha se saldo negativo
            if saldo_acumulado < 0:
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            
            data_atual += timedelta(days=1)
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        
        # Formatar números como moeda
        for row in range(6, ws.max_row + 1):
            for col in ['B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].number_format = 'R$ #,##0.00'
        
        # Salvar em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'fluxo_caixa_{data_hoje.strftime("%Y%m%d")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        flash(f'Erro ao exportar: {str(e)}', 'danger')
        return redirect(url_for('financeiro.fluxo_caixa'))


# ==================== ROTAS DE DRE ====================

@bp_financeiro.route('/dre')
def dre():
    """Demonstrativo de Resultados do Exercício."""
    try:
        from datetime import datetime
        from calendar import monthrange
        
        # Filtros
        ano = request.args.get('ano', type=int, default=date.today().year)
        mes = request.args.get('mes', type=int)
        comparacao = request.args.get('comparacao', 'mensal')  # mensal ou anual
        
        # Definir período
        if mes:
            # DRE mensal
            data_inicio = date(ano, mes, 1)
            ultimo_dia = monthrange(ano, mes)[1]
            data_fim = date(ano, mes, ultimo_dia)
            titulo_periodo = f"{data_inicio.strftime('%B/%Y').capitalize()}"
        else:
            # DRE anual
            data_inicio = date(ano, 1, 1)
            data_fim = date(ano, 12, 31)
            titulo_periodo = f"Ano {ano}"
        
        # Buscar lançamentos do período
        lancamentos = LancamentoFinanceiro.query.filter(
            LancamentoFinanceiro.data_lancamento >= data_inicio,
            LancamentoFinanceiro.data_lancamento <= data_fim,
            LancamentoFinanceiro.ativo == True
        ).all()
        
        # Calcular DRE
        dre_data = calcular_dre(lancamentos)
        
        # Comparação com período anterior
        if comparacao == 'mensal' and mes:
            # Mês anterior
            if mes == 1:
                mes_anterior = 12
                ano_anterior = ano - 1
            else:
                mes_anterior = mes - 1
                ano_anterior = ano
            
            data_inicio_ant = date(ano_anterior, mes_anterior, 1)
            ultimo_dia_ant = monthrange(ano_anterior, mes_anterior)[1]
            data_fim_ant = date(ano_anterior, mes_anterior, ultimo_dia_ant)
            
            lancamentos_ant = LancamentoFinanceiro.query.filter(
                LancamentoFinanceiro.data_lancamento >= data_inicio_ant,
                LancamentoFinanceiro.data_lancamento <= data_fim_ant,
                LancamentoFinanceiro.ativo == True
            ).all()
            
            dre_anterior = calcular_dre(lancamentos_ant)
            titulo_comparacao = f"{data_inicio_ant.strftime('%B/%Y').capitalize()}"
            
        elif comparacao == 'anual':
            # Ano anterior
            data_inicio_ant = date(ano - 1, 1, 1)
            data_fim_ant = date(ano - 1, 12, 31)
            
            lancamentos_ant = LancamentoFinanceiro.query.filter(
                LancamentoFinanceiro.data_lancamento >= data_inicio_ant,
                LancamentoFinanceiro.data_lancamento <= data_fim_ant,
                LancamentoFinanceiro.ativo == True
            ).all()
            
            dre_anterior = calcular_dre(lancamentos_ant)
            titulo_comparacao = f"Ano {ano - 1}"
        else:
            dre_anterior = None
            titulo_comparacao = None
        
        # Calcular variações
        variacoes = {}
        if dre_anterior:
            for key in dre_data.keys():
                valor_atual = dre_data[key]
                valor_anterior = dre_anterior[key]
                
                if valor_anterior != 0:
                    variacao_percentual = ((valor_atual - valor_anterior) / abs(valor_anterior)) * 100
                else:
                    variacao_percentual = 100 if valor_atual > 0 else 0
                
                variacoes[key] = {
                    'valor': valor_atual - valor_anterior,
                    'percentual': variacao_percentual
                }
        
        # DRE mensal (todos os meses do ano)
        dre_mensal = []
        if not mes:
            for m in range(1, 13):
                data_ini_mes = date(ano, m, 1)
                ultimo_dia_mes = monthrange(ano, m)[1]
                data_fim_mes = date(ano, m, ultimo_dia_mes)
                
                lanc_mes = LancamentoFinanceiro.query.filter(
                    LancamentoFinanceiro.data_lancamento >= data_ini_mes,
                    LancamentoFinanceiro.data_lancamento <= data_fim_mes,
                    LancamentoFinanceiro.ativo == True
                ).all()
                
                dre_mes = calcular_dre(lanc_mes)
                dre_mes['mes'] = m
                dre_mes['mes_nome'] = data_ini_mes.strftime('%B').capitalize()
                dre_mensal.append(dre_mes)
        
        return render_template('financeiro/dre/dashboard.html',
                             dre=dre_data,
                             dre_anterior=dre_anterior,
                             variacoes=variacoes,
                             dre_mensal=dre_mensal,
                             ano=ano,
                             mes=mes,
                             comparacao=comparacao,
                             titulo_periodo=titulo_periodo,
                             titulo_comparacao=titulo_comparacao)
    
    except Exception as e:
        flash(f'Erro ao gerar DRE: {str(e)}', 'danger')
        return redirect(url_for('financeiro.dashboard'))


def calcular_dre(lancamentos):
    """Calcula DRE a partir de uma lista de lançamentos."""
    
    # Categorias de receitas
    categorias_receita = ['Vendas', 'Serviços', 'Receitas Diversas', 'Receitas Financeiras']
    
    # Categorias de deduções
    categorias_deducoes = ['Impostos sobre Vendas', 'Devoluções', 'Descontos Concedidos']
    
    # Categorias de custos
    categorias_custos = ['Custo de Mercadorias', 'Custo de Serviços', 'Matéria-Prima']
    
    # Categorias de despesas operacionais
    categorias_despesas = ['Aluguel', 'Salários', 'Encargos', 'Energia', 'Água', 
                          'Internet', 'Telefone', 'Software', 'Manutenção', 
                          'Marketing', 'Contabilidade', 'Administrativas']
    
    # Categorias financeiras
    categorias_financeiras = ['Juros Pagos', 'Juros Recebidos', 'Despesas Bancárias']
    
    # Inicializar valores
    receita_bruta = Decimal('0.00')
    deducoes = Decimal('0.00')
    custos = Decimal('0.00')
    despesas_operacionais = Decimal('0.00')
    receitas_financeiras = Decimal('0.00')
    despesas_financeiras = Decimal('0.00')
    
    # Processar lançamentos
    for lanc in lancamentos:
        valor = lanc.valor
        categoria = lanc.categoria or ''
        
        if lanc.tipo == 'RECEITA':
            if categoria in categorias_deducoes:
                deducoes += valor
            elif categoria in categorias_financeiras or 'Juros Recebidos' in categoria:
                receitas_financeiras += valor
            else:
                receita_bruta += valor
        
        elif lanc.tipo == 'DESPESA':
            if categoria in categorias_custos:
                custos += valor
            elif categoria in categorias_financeiras or 'Juros' in categoria or 'Bancár' in categoria:
                despesas_financeiras += valor
            elif categoria in categorias_despesas or categoria in categorias_deducoes:
                if categoria in categorias_deducoes:
                    deducoes += valor
                else:
                    despesas_operacionais += valor
            else:
                despesas_operacionais += valor
    
    # Cálculos do DRE
    receita_liquida = receita_bruta - deducoes
    lucro_bruto = receita_liquida - custos
    lucro_operacional = lucro_bruto - despesas_operacionais
    resultado_financeiro = receitas_financeiras - despesas_financeiras
    lucro_liquido = lucro_operacional + resultado_financeiro
    
    # Margens (%)
    margem_bruta = (lucro_bruto / receita_liquida * 100) if receita_liquida > 0 else 0
    margem_operacional = (lucro_operacional / receita_liquida * 100) if receita_liquida > 0 else 0
    margem_liquida = (lucro_liquido / receita_liquida * 100) if receita_liquida > 0 else 0
    
    return {
        'receita_bruta': receita_bruta,
        'deducoes': deducoes,
        'receita_liquida': receita_liquida,
        'custos': custos,
        'lucro_bruto': lucro_bruto,
        'despesas_operacionais': despesas_operacionais,
        'lucro_operacional': lucro_operacional,
        'receitas_financeiras': receitas_financeiras,
        'despesas_financeiras': despesas_financeiras,
        'resultado_financeiro': resultado_financeiro,
        'lucro_liquido': lucro_liquido,
        'margem_bruta': margem_bruta,
        'margem_operacional': margem_operacional,
        'margem_liquida': margem_liquida
    }


@bp_financeiro.route('/dre/exportar-excel')
def exportar_dre_excel():
    """Exportar DRE para Excel."""
    try:
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from calendar import monthrange
        
        # Filtros
        ano = request.args.get('ano', type=int, default=date.today().year)
        mes = request.args.get('mes', type=int)
        
        # Definir período
        if mes:
            data_inicio = date(ano, mes, 1)
            ultimo_dia = monthrange(ano, mes)[1]
            data_fim = date(ano, mes, ultimo_dia)
            titulo_periodo = f"{data_inicio.strftime('%B/%Y').capitalize()}"
        else:
            data_inicio = date(ano, 1, 1)
            data_fim = date(ano, 12, 31)
            titulo_periodo = f"Ano {ano}"
        
        # Buscar lançamentos
        lancamentos = LancamentoFinanceiro.query.filter(
            LancamentoFinanceiro.data_lancamento >= data_inicio,
            LancamentoFinanceiro.data_lancamento <= data_fim,
            LancamentoFinanceiro.ativo == True
        ).all()
        
        dre_data = calcular_dre(lancamentos)
        
        # Criar Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DRE"
        
        # Estilos
        titulo_font = Font(size=14, bold=True)
        header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        subtotal_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        total_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        total_font = Font(color='FFFFFF', bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Cabeçalho
        ws['A1'] = 'DEMONSTRATIVO DE RESULTADOS DO EXERCÍCIO (DRE)'
        ws['A1'].font = titulo_font
        ws['A2'] = f'Período: {titulo_periodo}'
        ws.merge_cells('A1:C1')
        ws.merge_cells('A2:C2')
        ws.append([])
        
        # Colunas
        ws.append(['Descrição', 'Valor (R$)', '% s/ Rec. Líq.'])
        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Dados
        def add_linha(descricao, valor, percentual=None, estilo=None):
            row = [descricao, float(valor)]
            if percentual is not None:
                row.append(f"{percentual:.2f}%")
            else:
                row.append('')
            ws.append(row)
            
            if estilo == 'subtotal':
                for cell in ws[ws.max_row]:
                    cell.fill = subtotal_fill
                    cell.font = Font(bold=True)
                    cell.border = border
            elif estilo == 'total':
                for cell in ws[ws.max_row]:
                    cell.fill = total_fill
                    cell.font = total_font
                    cell.border = border
            else:
                for cell in ws[ws.max_row]:
                    cell.border = border
        
        # Receitas
        perc_rec_bruta = (dre_data['receita_bruta'] / dre_data['receita_liquida'] * 100) if dre_data['receita_liquida'] > 0 else 0
        add_linha('RECEITA BRUTA', dre_data['receita_bruta'], perc_rec_bruta)
        
        perc_deducoes = (dre_data['deducoes'] / dre_data['receita_liquida'] * 100) if dre_data['receita_liquida'] > 0 else 0
        add_linha('(-) Deduções', -dre_data['deducoes'], perc_deducoes)
        
        add_linha('RECEITA LÍQUIDA', dre_data['receita_liquida'], 100, 'subtotal')
        ws.append([])
        
        # Custos e Lucro Bruto
        perc_custos = (dre_data['custos'] / dre_data['receita_liquida'] * 100) if dre_data['receita_liquida'] > 0 else 0
        add_linha('(-) Custos', -dre_data['custos'], perc_custos)
        
        add_linha('LUCRO BRUTO', dre_data['lucro_bruto'], dre_data['margem_bruta'], 'subtotal')
        ws.append([])
        
        # Despesas Operacionais
        perc_desp = (dre_data['despesas_operacionais'] / dre_data['receita_liquida'] * 100) if dre_data['receita_liquida'] > 0 else 0
        add_linha('(-) Despesas Operacionais', -dre_data['despesas_operacionais'], perc_desp)
        
        add_linha('LUCRO OPERACIONAL', dre_data['lucro_operacional'], dre_data['margem_operacional'], 'subtotal')
        ws.append([])
        
        # Resultado Financeiro
        perc_rec_fin = (dre_data['receitas_financeiras'] / dre_data['receita_liquida'] * 100) if dre_data['receita_liquida'] > 0 else 0
        add_linha('(+) Receitas Financeiras', dre_data['receitas_financeiras'], perc_rec_fin)
        
        perc_desp_fin = (dre_data['despesas_financeiras'] / dre_data['receita_liquida'] * 100) if dre_data['receita_liquida'] > 0 else 0
        add_linha('(-) Despesas Financeiras', -dre_data['despesas_financeiras'], perc_desp_fin)
        
        add_linha('RESULTADO FINANCEIRO', dre_data['resultado_financeiro'], None, 'subtotal')
        ws.append([])
        
        # Lucro Líquido
        add_linha('LUCRO LÍQUIDO', dre_data['lucro_liquido'], dre_data['margem_liquida'], 'total')
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        
        # Formatar números
        for row in range(5, ws.max_row + 1):
            ws[f'B{row}'].number_format = 'R$ #,##0.00'
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            ws[f'C{row}'].alignment = Alignment(horizontal='center')
        
        # Salvar
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'dre_{titulo_periodo.replace("/", "-")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        flash(f'Erro ao exportar DRE: {str(e)}', 'danger')
        return redirect(url_for('financeiro.dre'))


# ============================================================
# PLANO DE CONTAS
# ============================================================

@bp_financeiro.route('/plano-contas')
def plano_contas():
    """Lista o plano de contas hierárquico."""
    from app.financeiro.financeiro_model import PlanoContas
    
    # Buscar todas as contas
    contas = PlanoContas.query.filter_by(ativa=True).order_by(PlanoContas.codigo).all()
    
    # Organizar por tipo
    contas_por_tipo = {
        'ATIVO': [],
        'PASSIVO': [],
        'RECEITA': [],
        'DESPESA': []
    }
    
    for conta in contas:
        if conta.tipo in contas_por_tipo:
            contas_por_tipo[conta.tipo].append(conta)
    
    return render_template('financeiro/plano_contas/listar.html',
                          contas_por_tipo=contas_por_tipo,
                          total_contas=len(contas))


@bp_financeiro.route('/plano-contas/nova', methods=['GET', 'POST'])
def plano_contas_nova():
    """Criar nova conta contábil."""
    from app.financeiro.financeiro_model import PlanoContas
    
    if request.method == 'POST':
        try:
            # Buscar conta pai se informada
            conta_pai_id = request.form.get('conta_pai_id')
            conta_pai = None
            nivel = 1
            
            if conta_pai_id:
                conta_pai = PlanoContas.query.get(conta_pai_id)
                if conta_pai:
                    nivel = conta_pai.nivel + 1
            
            # Criar conta
            conta = PlanoContas(
                codigo=request.form['codigo'],
                nome=request.form['nome'],
                descricao=request.form.get('descricao'),
                tipo=request.form['tipo'],
                nivel=nivel,
                conta_pai_id=conta_pai.id if conta_pai else None,
                aceita_lancamento=request.form.get('aceita_lancamento') == 'on',
                natureza=request.form.get('natureza'),
                ordem=int(request.form.get('ordem', 0))
            )
            
            db.session.add(conta)
            db.session.commit()
            
            flash(f'✅ Conta {conta.codigo} - {conta.nome} criada com sucesso!', 'success')
            return redirect(url_for('financeiro.plano_contas'))
        
        except Exception as e:
            db.session.rollback()
            flash(f'❌ Erro ao criar conta: {str(e)}', 'danger')
    
    # Buscar contas para serem pai
    contas_pai = PlanoContas.query.filter_by(ativa=True).order_by(PlanoContas.codigo).all()
    
    return render_template('financeiro/plano_contas/form.html',
                          contas_pai=contas_pai)


@bp_financeiro.route('/plano-contas/<int:id>/editar', methods=['GET', 'POST'])
def plano_contas_editar(id):
    """Editar conta contábil."""
    from app.financeiro.financeiro_model import PlanoContas
    
    conta = PlanoContas.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            conta.codigo = request.form['codigo']
            conta.nome = request.form['nome']
            conta.descricao = request.form.get('descricao')
            conta.tipo = request.form['tipo']
            conta.aceita_lancamento = request.form.get('aceita_lancamento') == 'on'
            conta.natureza = request.form.get('natureza')
            conta.ordem = int(request.form.get('ordem', 0))
            
            # Atualizar conta pai
            conta_pai_id = request.form.get('conta_pai_id')
            if conta_pai_id:
                conta_pai = PlanoContas.query.get(conta_pai_id)
                conta.conta_pai_id = conta_pai.id
                conta.nivel = conta_pai.nivel + 1
            else:
                conta.conta_pai_id = None
                conta.nivel = 1
            
            db.session.commit()
            
            flash(f'✅ Conta {conta.codigo} atualizada com sucesso!', 'success')
            return redirect(url_for('financeiro.plano_contas'))
        
        except Exception as e:
            db.session.rollback()
            flash(f'❌ Erro ao atualizar conta: {str(e)}', 'danger')
    
    # Buscar contas para serem pai (excluir a própria e suas filhas)
    contas_pai = PlanoContas.query.filter(
        PlanoContas.id != id,
        PlanoContas.ativa == True
    ).order_by(PlanoContas.codigo).all()
    
    return render_template('financeiro/plano_contas/form.html',
                          conta=conta,
                          contas_pai=contas_pai)


@bp_financeiro.route('/plano-contas/<int:id>/excluir', methods=['POST'])
def plano_contas_excluir(id):
    """Desativar conta contábil."""
    from app.financeiro.financeiro_model import PlanoContas
    
    try:
        conta = PlanoContas.query.get_or_404(id)
        
        # Verificar se tem lançamentos
        if len(conta.lancamentos) > 0:
            flash(f'❌ Não é possível excluir conta com lançamentos vinculados!', 'danger')
            return redirect(url_for('financeiro.plano_contas'))
        
        # Desativar
        conta.ativa = False
        db.session.commit()
        
        flash(f'✅ Conta {conta.codigo} desativada com sucesso!', 'success')
    
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Erro ao excluir conta: {str(e)}', 'danger')
    
    return redirect(url_for('financeiro.plano_contas'))


@bp_financeiro.route('/plano-contas/criar-padrao', methods=['POST'])
def plano_contas_criar_padrao():
    """Cria estrutura padrão de plano de contas."""
    from app.financeiro.financeiro_model import PlanoContas
    
    try:
        total = PlanoContas.criar_plano_padrao()
        flash(f'✅ {total} contas criadas no plano padrão!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Erro ao criar plano padrão: {str(e)}', 'danger')
    
    return redirect(url_for('financeiro.plano_contas'))


@bp_financeiro.route('/plano-contas/<int:id>/detalhes')
def plano_contas_detalhes(id):
    """Visualizar detalhes de uma conta."""
    from app.financeiro.financeiro_model import PlanoContas
    
    conta = PlanoContas.query.get_or_404(id)
    
    # Calcular saldo atual
    saldo_atual = conta.get_saldo()
    
    # Buscar últimos lançamentos
    lancamentos = LancamentoFinanceiro.query.filter_by(
        plano_conta_id=id
    ).order_by(LancamentoFinanceiro.data_lancamento.desc()).limit(20).all()
    
    return render_template('financeiro/plano_contas/detalhes.html',
                          conta=conta,
                          saldo_atual=saldo_atual,
                          lancamentos=lancamentos)


@bp_financeiro.route('/api/plano-contas/analiticas')
def api_plano_contas_analiticas():
    """API: Retorna apenas contas analíticas (que aceitam lançamento)."""
    from app.financeiro.financeiro_model import PlanoContas
    
    tipo = request.args.get('tipo')  # Filtrar por tipo se necessário
    
    query = PlanoContas.query.filter_by(ativa=True, aceita_lancamento=True)
    
    if tipo:
        query = query.filter_by(tipo=tipo)
    
    contas = query.order_by(PlanoContas.codigo).all()
    
    return jsonify({
        'success': True,
        'contas': [conta.to_dict() for conta in contas]
    })


# ============================================================================
# ORÇAMENTO ANUAL
# ============================================================================

@bp_financeiro.route('/orcamento-anual')
def orcamento_anual_listar():
    """Listar orçamentos anuais."""
    from app.financeiro.financeiro_model import OrcamentoAnual
    from datetime import datetime
    
    # Filtros
    ano = request.args.get('ano', datetime.now().year, type=int)
    tipo = request.args.get('tipo', '')
    categoria = request.args.get('categoria', '')
    
    # Query base
    query = OrcamentoAnual.query.filter_by(ano=ano)
    
    if tipo:
        query = query.filter_by(tipo=tipo)
    if categoria:
        query = query.filter(OrcamentoAnual.categoria.like(f'%{categoria}%'))
    
    orcamentos = query.order_by(OrcamentoAnual.mes, OrcamentoAnual.tipo, OrcamentoAnual.categoria).all()
    
    # Estatísticas
    total_orcado_receita = sum(o.valor_orcado for o in orcamentos if o.tipo == 'RECEITA')
    total_orcado_despesa = sum(o.valor_orcado for o in orcamentos if o.tipo == 'DESPESA')
    
    total_realizado_receita = sum(o.valor_realizado for o in orcamentos if o.tipo == 'RECEITA')
    total_realizado_despesa = sum(o.valor_realizado for o in orcamentos if o.tipo == 'DESPESA')
    
    # Anos disponíveis
    anos_disponiveis = db.session.query(
        db.func.distinct(OrcamentoAnual.ano)
    ).order_by(OrcamentoAnual.ano.desc()).all()
    anos_disponiveis = [a[0] for a in anos_disponiveis]
    
    # Categorias únicas
    categorias_disponiveis = db.session.query(
        db.func.distinct(OrcamentoAnual.categoria)
    ).filter_by(ano=ano).order_by(OrcamentoAnual.categoria).all()
    categorias_disponiveis = [c[0] for c in categorias_disponiveis]
    
    return render_template('financeiro/orcamento_anual/listar.html',
                         orcamentos=orcamentos,
                         ano_selecionado=ano,
                         tipo_selecionado=tipo,
                         categoria_selecionada=categoria,
                         anos_disponiveis=anos_disponiveis,
                         categorias_disponiveis=categorias_disponiveis,
                         total_orcado_receita=total_orcado_receita,
                         total_orcado_despesa=total_orcado_despesa,
                         total_realizado_receita=total_realizado_receita,
                         total_realizado_despesa=total_realizado_despesa)


@bp_financeiro.route('/orcamento-anual/dashboard')
def orcamento_anual_dashboard():
    """Dashboard analítico do orçamento anual."""
    from app.financeiro.financeiro_model import OrcamentoAnual
    from datetime import datetime
    
    ano = request.args.get('ano', datetime.now().year, type=int)
    
    # Busca todos os orçamentos do ano
    orcamentos = OrcamentoAnual.query.filter_by(ano=ano, ativo=True).all()
    
    # Dados para gráficos
    dados_mensais = {}
    for mes in range(1, 13):
        orcamentos_mes = [o for o in orcamentos if o.mes == mes]
        
        dados_mensais[mes] = {
            'mes': mes,
            'mes_nome': OrcamentoAnual(mes=mes, ano=ano, tipo='RECEITA', categoria='temp', valor_orcado=0).mes_nome,
            'orcado_receita': sum(o.valor_orcado for o in orcamentos_mes if o.tipo == 'RECEITA'),
            'orcado_despesa': sum(o.valor_orcado for o in orcamentos_mes if o.tipo == 'DESPESA'),
            'realizado_receita': sum(o.valor_realizado for o in orcamentos_mes if o.tipo == 'RECEITA'),
            'realizado_despesa': sum(o.valor_realizado for o in orcamentos_mes if o.tipo == 'DESPESA'),
        }
    
    # Análise por categoria
    categorias = {}
    for orc in orcamentos:
        if orc.categoria not in categorias:
            categorias[orc.categoria] = {
                'categoria': orc.categoria,
                'tipo': orc.tipo,
                'orcado': 0,
                'realizado': 0
            }
        categorias[orc.categoria]['orcado'] += float(orc.valor_orcado)
        categorias[orc.categoria]['realizado'] += orc.valor_realizado
    
    # Top 5 categorias por valor
    top_categorias = sorted(
        categorias.values(),
        key=lambda x: x['orcado'],
        reverse=True
    )[:10]
    
    # Anos disponíveis
    anos_disponiveis = db.session.query(
        db.func.distinct(OrcamentoAnual.ano)
    ).order_by(OrcamentoAnual.ano.desc()).all()
    anos_disponiveis = [a[0] for a in anos_disponiveis]
    
    # Indicadores
    total_orcado_receita = sum(c['orcado'] for c in categorias.values() if c['tipo'] == 'RECEITA')
    total_orcado_despesa = sum(c['orcado'] for c in categorias.values() if c['tipo'] == 'DESPESA')
    total_realizado_receita = sum(c['realizado'] for c in categorias.values() if c['tipo'] == 'RECEITA')
    total_realizado_despesa = sum(c['realizado'] for c in categorias.values() if c['tipo'] == 'DESPESA')
    
    resultado_orcado = total_orcado_receita - total_orcado_despesa
    resultado_realizado = total_realizado_receita - total_realizado_despesa
    
    return render_template('financeiro/orcamento_anual/dashboard.html',
                         ano_selecionado=ano,
                         anos_disponiveis=anos_disponiveis,
                         dados_mensais=list(dados_mensais.values()),
                         top_categorias=top_categorias,
                         total_orcado_receita=total_orcado_receita,
                         total_orcado_despesa=total_orcado_despesa,
                         total_realizado_receita=total_realizado_receita,
                         total_realizado_despesa=total_realizado_despesa,
                         resultado_orcado=resultado_orcado,
                         resultado_realizado=resultado_realizado)


@bp_financeiro.route('/orcamento-anual/novo', methods=['GET', 'POST'])
def orcamento_anual_novo():
    """Criar novo orçamento."""
    from app.financeiro.financeiro_model import OrcamentoAnual, CentroCusto, PlanoContas
    from datetime import datetime
    
    if request.method == 'POST':
        try:
            # Validações
            ano = int(request.form['ano'])
            mes = int(request.form['mes'])
            tipo = request.form['tipo']
            categoria = request.form['categoria']
            valor_orcado = float(request.form['valor_orcado'].replace(',', '.'))
            
            # Verifica duplicidade
            existe = OrcamentoAnual.query.filter_by(
                ano=ano,
                mes=mes,
                tipo=tipo,
                categoria=categoria
            ).first()
            
            if existe:
                flash('Já existe um orçamento para esta combinação de ano/mês/tipo/categoria!', 'warning')
                return redirect(url_for('financeiro.orcamento_anual_novo'))
            
            # Cria orçamento
            orcamento = OrcamentoAnual(
                ano=ano,
                mes=mes,
                tipo=tipo,
                categoria=categoria,
                valor_orcado=valor_orcado,
                descricao=request.form.get('descricao', ''),
                observacoes=request.form.get('observacoes', ''),
                centro_custo_id=request.form.get('centro_custo_id') or None,
                plano_conta_id=request.form.get('plano_conta_id') or None,
                ativo=True
            )
            
            db.session.add(orcamento)
            db.session.commit()
            
            flash('Orçamento criado com sucesso!', 'success')
            return redirect(url_for('financeiro.orcamento_anual_listar', ano=ano))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar orçamento: {str(e)}', 'danger')
    
    # GET
    centros_custo = CentroCusto.query.filter_by(ativo=True).order_by(CentroCusto.nome).all()
    contas = PlanoContas.query.filter_by(ativa=True, aceita_lancamento=True).order_by(PlanoContas.codigo).all()
    
    ano_atual = datetime.now().year
    
    return render_template('financeiro/orcamento_anual/form.html',
                         centros_custo=centros_custo,
                         contas=contas,
                         ano_atual=ano_atual)


@bp_financeiro.route('/orcamento-anual/<int:id>/editar', methods=['GET', 'POST'])
def orcamento_anual_editar(id):
    """Editar orçamento."""
    from app.financeiro.financeiro_model import OrcamentoAnual, CentroCusto, PlanoContas
    
    orcamento = OrcamentoAnual.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            orcamento.ano = int(request.form['ano'])
            orcamento.mes = int(request.form['mes'])
            orcamento.tipo = request.form['tipo']
            orcamento.categoria = request.form['categoria']
            orcamento.valor_orcado = float(request.form['valor_orcado'].replace(',', '.'))
            orcamento.descricao = request.form.get('descricao', '')
            orcamento.observacoes = request.form.get('observacoes', '')
            orcamento.centro_custo_id = request.form.get('centro_custo_id') or None
            orcamento.plano_conta_id = request.form.get('plano_conta_id') or None
            orcamento.ativo = 'ativo' in request.form
            
            db.session.commit()
            
            flash('Orçamento atualizado com sucesso!', 'success')
            return redirect(url_for('financeiro.orcamento_anual_listar', ano=orcamento.ano))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar orçamento: {str(e)}', 'danger')
    
    # GET
    centros_custo = CentroCusto.query.filter_by(ativo=True).order_by(CentroCusto.nome).all()
    contas = PlanoContas.query.filter_by(ativa=True, aceita_lancamento=True).order_by(PlanoContas.codigo).all()
    
    return render_template('financeiro/orcamento_anual/form.html',
                         orcamento=orcamento,
                         centros_custo=centros_custo,
                         contas=contas)


@bp_financeiro.route('/orcamento-anual/<int:id>/excluir', methods=['POST'])
def orcamento_anual_excluir(id):
    """Excluir orçamento."""
    from app.financeiro.financeiro_model import OrcamentoAnual
    
    orcamento = OrcamentoAnual.query.get_or_404(id)
    ano = orcamento.ano
    
    try:
        db.session.delete(orcamento)
        db.session.commit()
        flash('Orçamento excluído com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir orçamento: {str(e)}', 'danger')
    
    return redirect(url_for('financeiro.orcamento_anual_listar', ano=ano))


@bp_financeiro.route('/orcamento-anual/criar-padrao', methods=['POST'])
def orcamento_anual_criar_padrao():
    """Criar orçamentos padrão para um ano."""
    from app.financeiro.financeiro_model import OrcamentoAnual
    from datetime import datetime
    
    ano = request.form.get('ano', datetime.now().year, type=int)
    
    try:
        qtd = OrcamentoAnual.criar_orcamento_padrao(ano)
        flash(f'✅ {qtd} orçamentos padrão criados para {ano}!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao criar orçamentos padrão: {str(e)}', 'danger')
    
    return redirect(url_for('financeiro.orcamento_anual_listar', ano=ano))


@bp_financeiro.route('/orcamento-anual/comparacao')
def orcamento_anual_comparacao():
    """Comparação Orçado vs Realizado."""
    from app.financeiro.financeiro_model import OrcamentoAnual
    from datetime import datetime
    
    ano = request.args.get('ano', datetime.now().year, type=int)
    mes = request.args.get('mes', type=int)
    
    query = OrcamentoAnual.query.filter_by(ano=ano, ativo=True)
    
    if mes:
        query = query.filter_by(mes=mes)
    
    orcamentos = query.order_by(
        OrcamentoAnual.mes,
        OrcamentoAnual.tipo,
        OrcamentoAnual.categoria
    ).all()
    
    # Agrupa por categoria
    categorias = {}
    for orc in orcamentos:
        key = f"{orc.tipo}_{orc.categoria}"
        if key not in categorias:
            categorias[key] = {
                'tipo': orc.tipo,
                'categoria': orc.categoria,
                'orcado': 0,
                'realizado': 0,
                'meses': []
            }
        categorias[key]['orcado'] += float(orc.valor_orcado)
        categorias[key]['realizado'] += orc.valor_realizado
        categorias[key]['meses'].append({
            'mes': orc.mes,
            'mes_nome': orc.mes_nome,
            'orcado': float(orc.valor_orcado),
            'realizado': orc.valor_realizado,
            'percentual': orc.percentual_executado
        })
    
    # Converte para lista
    dados_comparacao = list(categorias.values())
    
    # Anos disponíveis
    anos_disponiveis = db.session.query(
        db.func.distinct(OrcamentoAnual.ano)
    ).order_by(OrcamentoAnual.ano.desc()).all()
    anos_disponiveis = [a[0] for a in anos_disponiveis]
    
    return render_template('financeiro/orcamento_anual/comparacao.html',
                         dados_comparacao=dados_comparacao,
                         ano_selecionado=ano,
                         mes_selecionado=mes,
                         anos_disponiveis=anos_disponiveis)


# ============================================================================
# NOTAS FISCAIS
# ============================================================================

@bp_financeiro.route('/notas-fiscais')
def notas_fiscais_listar():
    """Listar notas fiscais."""
    from app.financeiro.financeiro_model import NotaFiscal
    from datetime import datetime, timedelta
    
    # Filtros
    tipo = request.args.get('tipo', '')
    status = request.args.get('status', '')
    data_inicio = request.args.get('data_inicio', '')
    data_fim = request.args.get('data_fim', '')
    busca = request.args.get('busca', '')
    
    # Query base
    query = NotaFiscal.query
    
    if tipo:
        query = query.filter_by(tipo=tipo)
    if status:
        query = query.filter_by(status=status)
    if data_inicio:
        query = query.filter(NotaFiscal.data_emissao >= datetime.strptime(data_inicio, '%Y-%m-%d').date())
    if data_fim:
        query = query.filter(NotaFiscal.data_emissao <= datetime.strptime(data_fim, '%Y-%m-%d').date())
    if busca:
        query = query.filter(
            db.or_(
                NotaFiscal.numero.like(f'%{busca}%'),
                NotaFiscal.chave_acesso.like(f'%{busca}%'),
                NotaFiscal.emitente_nome.like(f'%{busca}%'),
                NotaFiscal.destinatario_nome.like(f'%{busca}%')
            )
        )
    
    notas = query.order_by(NotaFiscal.data_emissao.desc()).all()
    
    # Estatísticas
    total_valor = sum(n.valor_total for n in notas)
    total_entrada = sum(n.valor_total for n in notas if n.tipo == 'ENTRADA')
    total_saida = sum(n.valor_total for n in notas if n.tipo == 'SAIDA')
    
    # Status count
    status_count = {
        'PENDENTE': len([n for n in notas if n.status == 'PENDENTE']),
        'PROCESSADA': len([n for n in notas if n.status == 'PROCESSADA']),
        'PAGA': len([n for n in notas if n.status == 'PAGA']),
        'CANCELADA': len([n for n in notas if n.status == 'CANCELADA']),
    }
    
    return render_template('financeiro/notas_fiscais/listar.html',
                         notas=notas,
                         tipo_selecionado=tipo,
                         status_selecionado=status,
                         data_inicio=data_inicio,
                         data_fim=data_fim,
                         busca=busca,
                         total_valor=total_valor,
                         total_entrada=total_entrada,
                         total_saida=total_saida,
                         status_count=status_count)


@bp_financeiro.route('/notas-fiscais/nova', methods=['GET', 'POST'])
def notas_fiscais_nova():
    """Criar nova nota fiscal."""
    from app.financeiro.financeiro_model import NotaFiscal
    from app.cliente.cliente_model import Cliente
    from app.fornecedor.fornecedor_model import Fornecedor
    from datetime import datetime
    import base64
    
    if request.method == 'POST':
        try:
            # Upload de XML
            arquivo_xml = request.files.get('arquivo_xml')
            arquivo_pdf = request.files.get('arquivo_pdf')
            
            dados_nota = {}
            
            # Parse XML se fornecido
            if arquivo_xml and arquivo_xml.filename:
                xml_content = arquivo_xml.read()
                
                # Tenta fazer parse
                dados_xml = NotaFiscal.parse_xml_nfe(xml_content.decode('utf-8'))
                
                if 'error' in dados_xml:
                    flash(f'Erro ao processar XML: {dados_xml["error"]}', 'warning')
                else:
                    dados_nota = dados_xml
                    # Salva XML em base64
                    dados_nota['arquivo_xml'] = base64.b64encode(xml_content).decode('utf-8')
                    dados_nota['arquivo_xml_nome'] = arquivo_xml.filename
                    
                    # Determina tipo baseado no CNPJ
                    # Se destinatário é a empresa, é ENTRADA
                    # Aqui você pode ajustar com o CNPJ da sua empresa
                    dados_nota['tipo'] = request.form.get('tipo', 'ENTRADA')
            
            # Sobrescreve com dados do formulário (manual)
            nota = NotaFiscal(
                numero=request.form.get('numero') or dados_nota.get('numero'),
                serie=request.form.get('serie') or dados_nota.get('serie'),
                modelo=request.form.get('modelo') or dados_nota.get('modelo', '55'),
                tipo=request.form.get('tipo'),
                chave_acesso=request.form.get('chave_acesso') or dados_nota.get('chave_acesso'),
                data_emissao=datetime.strptime(request.form.get('data_emissao'), '%Y-%m-%d').date() if request.form.get('data_emissao') else dados_nota.get('data_emissao'),
                valor_total=float(request.form.get('valor_total', '0').replace(',', '.')) if request.form.get('valor_total') else dados_nota.get('valor_total', 0),
                valor_produtos=float(request.form.get('valor_produtos', '0').replace(',', '.')) if request.form.get('valor_produtos') else dados_nota.get('valor_produtos', 0),
                natureza_operacao=request.form.get('natureza_operacao') or dados_nota.get('natureza_operacao'),
                cfop=request.form.get('cfop') or dados_nota.get('cfop'),
                status=request.form.get('status', 'PENDENTE'),
                observacoes=request.form.get('observacoes', ''),
                cliente_id=request.form.get('cliente_id') or None,
                fornecedor_id=request.form.get('fornecedor_id') or None,
                arquivo_xml=dados_nota.get('arquivo_xml'),
                arquivo_xml_nome=dados_nota.get('arquivo_xml_nome'),
            )
            
            # Dados de emitente/destinatário
            if dados_nota:
                nota.emitente_nome = dados_nota.get('emitente_nome')
                nota.emitente_cnpj = dados_nota.get('emitente_cnpj')
                nota.destinatario_nome = dados_nota.get('destinatario_nome')
                nota.destinatario_cnpj = dados_nota.get('destinatario_cnpj')
                nota.valor_frete = dados_nota.get('valor_frete', 0)
                nota.valor_desconto = dados_nota.get('valor_desconto', 0)
                nota.valor_icms = dados_nota.get('valor_icms', 0)
                nota.valor_pis = dados_nota.get('valor_pis', 0)
                nota.valor_cofins = dados_nota.get('valor_cofins', 0)
            
            # Upload de PDF
            if arquivo_pdf and arquivo_pdf.filename:
                pdf_content = arquivo_pdf.read()
                nota.arquivo_pdf = base64.b64encode(pdf_content).decode('utf-8')
                nota.arquivo_pdf_nome = arquivo_pdf.filename
            
            db.session.add(nota)
            db.session.commit()
            
            # Criar lançamento automático se solicitado
            if request.form.get('criar_lancamento') == 'on':
                nota.criar_lancamento_automatico()
                db.session.commit()
                flash('Nota fiscal criada e lançamento financeiro gerado!', 'success')
            else:
                flash('Nota fiscal criada com sucesso!', 'success')
            
            return redirect(url_for('financeiro.notas_fiscais_visualizar', id=nota.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar nota fiscal: {str(e)}', 'danger')
            import traceback
            print(traceback.format_exc())
    
    # GET
    clientes = Cliente.query.filter_by(ativo=True).order_by(Cliente.nome).all()
    fornecedores = Fornecedor.query.filter_by(ativo=True).order_by(Fornecedor.nome).all()
    
    return render_template('financeiro/notas_fiscais/form.html',
                         clientes=clientes,
                         fornecedores=fornecedores)


@bp_financeiro.route('/notas-fiscais/<int:id>')
def notas_fiscais_visualizar(id):
    """Visualizar nota fiscal."""
    from app.financeiro.financeiro_model import NotaFiscal
    
    nota = NotaFiscal.query.get_or_404(id)
    
    return render_template('financeiro/notas_fiscais/visualizar.html',
                         nota=nota)


@bp_financeiro.route('/notas-fiscais/<int:id>/editar', methods=['GET', 'POST'])
def notas_fiscais_editar(id):
    """Editar nota fiscal."""
    from app.financeiro.financeiro_model import NotaFiscal
    from app.cliente.cliente_model import Cliente
    from app.fornecedor.fornecedor_model import Fornecedor
    from datetime import datetime
    
    nota = NotaFiscal.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            nota.numero = request.form['numero']
            nota.serie = request.form.get('serie', '')
            nota.modelo = request.form.get('modelo', '55')
            nota.tipo = request.form['tipo']
            nota.data_emissao = datetime.strptime(request.form['data_emissao'], '%Y-%m-%d').date()
            nota.valor_total = float(request.form['valor_total'].replace(',', '.'))
            nota.natureza_operacao = request.form.get('natureza_operacao', '')
            nota.status = request.form['status']
            nota.observacoes = request.form.get('observacoes', '')
            nota.cliente_id = request.form.get('cliente_id') or None
            nota.fornecedor_id = request.form.get('fornecedor_id') or None
            
            db.session.commit()
            flash('Nota fiscal atualizada com sucesso!', 'success')
            return redirect(url_for('financeiro.notas_fiscais_visualizar', id=nota.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar nota: {str(e)}', 'danger')
    
    # GET
    clientes = Cliente.query.filter_by(ativo=True).order_by(Cliente.nome).all()
    fornecedores = Fornecedor.query.filter_by(ativo=True).order_by(Fornecedor.nome).all()
    
    return render_template('financeiro/notas_fiscais/form.html',
                         nota=nota,
                         clientes=clientes,
                         fornecedores=fornecedores)


@bp_financeiro.route('/notas-fiscais/<int:id>/excluir', methods=['POST'])
def notas_fiscais_excluir(id):
    """Excluir nota fiscal."""
    from app.financeiro.financeiro_model import NotaFiscal
    
    nota = NotaFiscal.query.get_or_404(id)
    
    try:
        # Verifica se tem lançamento vinculado
        if nota.lancamento_id:
            flash('Não é possível excluir nota com lançamento financeiro vinculado!', 'danger')
            return redirect(url_for('financeiro.notas_fiscais_visualizar', id=id))
        
        db.session.delete(nota)
        db.session.commit()
        flash('Nota fiscal excluída com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir nota: {str(e)}', 'danger')
    
    return redirect(url_for('financeiro.notas_fiscais_listar'))


@bp_financeiro.route('/notas-fiscais/<int:id>/criar-lancamento', methods=['POST'])
def notas_fiscais_criar_lancamento(id):
    """Criar lançamento financeiro da nota."""
    from app.financeiro.financeiro_model import NotaFiscal
    
    nota = NotaFiscal.query.get_or_404(id)
    
    try:
        if nota.lancamento_id:
            flash('Esta nota já possui lançamento financeiro!', 'warning')
        else:
            lancamento = nota.criar_lancamento_automatico()
            db.session.commit()
            flash(f'Lançamento financeiro criado com sucesso! (#{lancamento.id})', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao criar lançamento: {str(e)}', 'danger')
    
    return redirect(url_for('financeiro.notas_fiscais_visualizar', id=id))


@bp_financeiro.route('/notas-fiscais/<int:id>/download/<tipo>')
def notas_fiscais_download(id, tipo):
    """Download de arquivo XML ou PDF."""
    from app.financeiro.financeiro_model import NotaFiscal
    import base64
    
    nota = NotaFiscal.query.get_or_404(id)
    
    try:
        if tipo == 'xml':
            if not nota.arquivo_xml:
                flash('Nota não possui arquivo XML!', 'warning')
                return redirect(url_for('financeiro.notas_fiscais_visualizar', id=id))
            
            arquivo_bytes = base64.b64decode(nota.arquivo_xml)
            filename = nota.arquivo_xml_nome or f'NF_{nota.numero_completo}.xml'
            
            return send_file(
                io.BytesIO(arquivo_bytes),
                mimetype='application/xml',
                as_attachment=True,
                download_name=filename
            )
        
        elif tipo == 'pdf':
            if not nota.arquivo_pdf:
                flash('Nota não possui arquivo PDF!', 'warning')
                return redirect(url_for('financeiro.notas_fiscais_visualizar', id=id))
            
            arquivo_bytes = base64.b64decode(nota.arquivo_pdf)
            filename = nota.arquivo_pdf_nome or f'NF_{nota.numero_completo}.pdf'
            
            return send_file(
                io.BytesIO(arquivo_bytes),
                mimetype='application/pdf',
                as_attachment=True,
                download_name=filename
            )
        
    except Exception as e:
        flash(f'Erro ao fazer download: {str(e)}', 'danger')
        return redirect(url_for('financeiro.notas_fiscais_visualizar', id=id))


@bp_financeiro.route('/notas-fiscais/galeria')
def notas_fiscais_galeria():
    """Galeria de notas fiscais com thumbnails."""
    from app.financeiro.financeiro_model import NotaFiscal
    from datetime import datetime, timedelta
    
    # Filtros
    mes = request.args.get('mes', type=int)
    ano = request.args.get('ano', datetime.now().year, type=int)
    tipo = request.args.get('tipo', '')
    
    query = NotaFiscal.query
    
    if mes:
        data_inicio = datetime(ano, mes, 1).date()
        if mes == 12:
            data_fim = datetime(ano + 1, 1, 1).date()
        else:
            data_fim = datetime(ano, mes + 1, 1).date()
        
        query = query.filter(
            NotaFiscal.data_emissao >= data_inicio,
            NotaFiscal.data_emissao < data_fim
        )
    else:
        query = query.filter(db.extract('year', NotaFiscal.data_emissao) == ano)
    
    if tipo:
        query = query.filter_by(tipo=tipo)
    
    notas = query.order_by(NotaFiscal.data_emissao.desc()).all()
    
    return render_template('financeiro/notas_fiscais/galeria.html',
                         notas=notas,
                         mes_selecionado=mes,
                         ano_selecionado=ano,
                         tipo_selecionado=tipo)


# ============================================================
# NOTIFICAÇÕES E ALERTAS
# ============================================================

@bp_financeiro.route('/notificacoes')
def notificacoes():
    """Lista todas as notificações."""
    try:
        from app.financeiro.financeiro_model import Notificacao
        
        # Filtros
        tipo = request.args.get('tipo')
        prioridade = request.args.get('prioridade')
        lidas = request.args.get('lidas', 'false') == 'true'
        
        # Query base
        query = Notificacao.query.filter_by(ativo=True)
        
        # Aplicar filtros
        if tipo:
            query = query.filter_by(tipo=tipo)
        
        if prioridade:
            query = query.filter_by(prioridade=prioridade)
        
        if not lidas:
            query = query.filter_by(lida=False)
        
        # Ordenar por data (mais recentes primeiro)
        notificacoes_lista = query.order_by(Notificacao.data_criacao.desc()).all()
        
        # Contar não lidas
        total_nao_lidas = Notificacao.contar_nao_lidas()
        
        return render_template('financeiro/notificacoes/listar.html',
                             notificacoes=notificacoes_lista,
                             total_nao_lidas=total_nao_lidas,
                             filtro_tipo=tipo,
                             filtro_prioridade=prioridade,
                             filtro_lidas=lidas)
    
    except Exception as e:
        flash(f'Erro ao carregar notificações: {str(e)}', 'danger')
        return redirect(url_for('financeiro.dashboard'))


@bp_financeiro.route('/notificacoes/<int:id>/marcar-lida', methods=['POST'])
def marcar_notificacao_lida(id):
    """Marca notificação como lida."""
    try:
        from app.financeiro.financeiro_model import Notificacao
        
        notificacao = Notificacao.query.get_or_404(id)
        notificacao.marcar_como_lida()
        
        return jsonify({'success': True, 'message': 'Notificação marcada como lida'})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@bp_financeiro.route('/notificacoes/marcar-todas-lidas', methods=['POST'])
def marcar_todas_lidas():
    """Marca todas as notificações como lidas."""
    try:
        from app.financeiro.financeiro_model import Notificacao
        
        notificacoes = Notificacao.get_nao_lidas()
        
        for notif in notificacoes:
            notif.marcar_como_lida()
        
        flash('Todas as notificações foram marcadas como lidas!', 'success')
        return redirect(url_for('financeiro.notificacoes'))
    
    except Exception as e:
        flash(f'Erro: {str(e)}', 'danger')
        return redirect(url_for('financeiro.notificacoes'))


@bp_financeiro.route('/notificacoes/verificar-alertas', methods=['POST'])
def verificar_alertas():
    """Executa verificação manual de alertas."""
    try:
        from app.financeiro.financeiro_model import Notificacao
        
        notificacoes = Notificacao.verificar_todas()
        
        flash(f'{len(notificacoes)} novas notificações criadas!', 'success')
        return redirect(url_for('financeiro.notificacoes'))
    
    except Exception as e:
        flash(f'Erro ao verificar alertas: {str(e)}', 'danger')
        return redirect(url_for('financeiro.notificacoes'))


@bp_financeiro.route('/api/notificacoes/nao-lidas')
def api_notificacoes_nao_lidas():
    """API para contar notificações não lidas (para badge no menu)."""
    try:
        from app.financeiro.financeiro_model import Notificacao
        
        total = Notificacao.contar_nao_lidas()
        
        return jsonify({
            'total': total,
            'success': True
        })
    
    except Exception as e:
        return jsonify({
            'error': str(e),
            'success': False
        }), 500


# ============================================================
# IMPORTAÇÃO EM LOTE
# ============================================================

@bp_financeiro.route('/importacao-lote')
def importacao_lote():
    """Lista importações realizadas."""
    try:
        from app.financeiro.financeiro_model import ImportacaoLote
        
        importacoes = ImportacaoLote.query.filter_by(ativo=True).order_by(
            ImportacaoLote.data_inicio.desc()
        ).all()
        
        return render_template('financeiro/importacao_lote/listar.html',
                             importacoes=importacoes)
    
    except Exception as e:
        flash(f'Erro: {str(e)}', 'danger')
        return redirect(url_for('financeiro.dashboard'))


@bp_financeiro.route('/importacao-lote/nova', methods=['GET', 'POST'])
def nova_importacao_lote():
    """Upload de arquivo para importação."""
    if request.method == 'POST':
        try:
            from app.financeiro.financeiro_model import ImportacaoLote
            import pandas as pd
            import os
            from werkzeug.utils import secure_filename
            
            # Verificar arquivo
            if 'arquivo' not in request.files:
                flash('Nenhum arquivo selecionado', 'danger')
                return redirect(request.url)
            
            arquivo = request.files['arquivo']
            
            if arquivo.filename == '':
                flash('Nenhum arquivo selecionado', 'danger')
                return redirect(request.url)
            
            # Salvar arquivo temporariamente
            filename = secure_filename(arquivo.filename)
            upload_folder = os.path.join('app', 'static', 'uploads', 'importacoes')
            os.makedirs(upload_folder, exist_ok=True)
            
            filepath = os.path.join(upload_folder, filename)
            arquivo.save(filepath)
            
            # Detectar tipo
            tipo_arquivo = 'EXCEL' if filename.endswith(('.xlsx', '.xls')) else 'CSV'
            
            # Criar registro de importação
            importacao = ImportacaoLote(
                arquivo_nome=filename,
                arquivo_path=filepath,
                tipo_arquivo=tipo_arquivo,
                status='PROCESSANDO'
            )
            db.session.add(importacao)
            db.session.commit()
            
            # Processar arquivo
            try:
                if tipo_arquivo == 'EXCEL':
                    df = pd.read_excel(filepath)
                else:
                    df = pd.read_csv(filepath)
                
                importacao.total_linhas = len(df)
                
                # Obter mapeamento de colunas do form
                mapa = {
                    'data': request.form.get('col_data'),
                    'descricao': request.form.get('col_descricao'),
                    'valor': request.form.get('col_valor'),
                    'tipo': request.form.get('col_tipo'),
                }
                
                erros = []
                sucesso = 0
                
                # Importar cada linha
                for idx, row in df.iterrows():
                    try:
                        # Converter data
                        data_lanc = pd.to_datetime(row[mapa['data']]).date()
                        
                        # Criar lançamento
                        lancamento = LancamentoFinanceiro(
                            descricao=str(row[mapa['descricao']]),
                            valor=converter_valor_monetario(str(row[mapa['valor']])),
                            tipo=str(row[mapa['tipo']]).upper(),
                            data_lancamento=data_lanc,
                            data_vencimento=data_lanc,
                            status='pendente'
                        )
                        
                        db.session.add(lancamento)
                        sucesso += 1
                    
                    except Exception as e:
                        erros.append({
                            'linha': idx + 2,  # +2 porque linha 1 é cabeçalho e idx começa em 0
                            'erro': str(e)
                        })
                
                # Salvar
                db.session.commit()
                
                # Atualizar importação
                importacao.linhas_importadas = sucesso
                importacao.linhas_erro = len(erros)
                
                if len(erros) > 0:
                    import json
                    importacao.erros_detalhes = json.dumps(erros)
                    importacao.status = 'PARCIAL' if sucesso > 0 else 'ERRO'
                else:
                    importacao.status = 'CONCLUIDA'
                
                importacao.finalizar(importacao.status)
                
                flash(f'Importação concluída! {sucesso} lançamentos importados, {len(erros)} erros.', 
                      'success' if len(erros) == 0 else 'warning')
                
                return redirect(url_for('financeiro.importacao_lote'))
            
            except Exception as e:
                importacao.status = 'ERRO'
                importacao.erros_detalhes = str(e)
                importacao.finalizar('ERRO')
                
                flash(f'Erro ao processar arquivo: {str(e)}', 'danger')
                return redirect(url_for('financeiro.importacao_lote'))
        
        except Exception as e:
            flash(f'Erro: {str(e)}', 'danger')
            return redirect(url_for('financeiro.importacao_lote'))
    
    # GET - Mostrar formulário
    return render_template('financeiro/importacao_lote/form.html')


@bp_financeiro.route('/importacao-lote/<int:id>/detalhes')
def detalhes_importacao(id):
    """Detalhes de uma importação."""
    try:
        from app.financeiro.financeiro_model import ImportacaoLote
        import json
        
        importacao = ImportacaoLote.query.get_or_404(id)
        
        # Parse dos erros
        erros = []
        if importacao.erros_detalhes:
            erros = json.loads(importacao.erros_detalhes)
        
        return render_template('financeiro/importacao_lote/detalhes.html',
                             importacao=importacao,
                             erros=erros)
    
    except Exception as e:
        flash(f'Erro: {str(e)}', 'danger')
        return redirect(url_for('financeiro.importacao_lote'))


# ============================================================
# RATEIO DE DESPESAS
# ============================================================

@bp_financeiro.route('/lancamentos/<int:id>/ratear', methods=['GET', 'POST'])
def ratear_despesa(id):
    """Rateio de despesa entre centros de custo."""
    lancamento = LancamentoFinanceiro.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            from app.financeiro.financeiro_model import RateioDespesa
            
            # Obter distribuição do formulário
            distribuicao = []
            
            # Contar quantos centros foram selecionados
            i = 0
            while True:
                centro_id = request.form.get(f'centro_custo_id_{i}')
                if not centro_id:
                    break
                
                percentual = request.form.get(f'percentual_{i}')
                observacoes = request.form.get(f'observacoes_{i}')
                
                distribuicao.append({
                    'centro_custo_id': int(centro_id),
                    'percentual': float(percentual),
                    'observacoes': observacoes
                })
                
                i += 1
            
            # Criar rateio
            RateioDespesa.criar_rateio(id, distribuicao)
            
            flash('Rateio criado com sucesso!', 'success')
            return redirect(url_for('financeiro.listar_lancamentos'))
        
        except ValueError as e:
            flash(str(e), 'danger')
            return redirect(request.url)
        
        except Exception as e:
            flash(f'Erro ao criar rateio: {str(e)}', 'danger')
            return redirect(request.url)
    
    # GET - Mostrar formulário
    from app.financeiro.financeiro_model import RateioDespesa
    
    centros_custo = CentroCusto.query.filter_by(ativo=True).all()
    rateios_existentes = RateioDespesa.get_por_lancamento(id)
    
    return render_template('financeiro/rateio/form.html',
                         lancamento=lancamento,
                         centros_custo=centros_custo,
                         rateios_existentes=rateios_existentes)


@bp_financeiro.route('/centros-custo/<int:id>/rateios')
def rateios_centro_custo(id):
    """Lista rateios recebidos por um centro de custo."""
    try:
        from app.financeiro.financeiro_model import RateioDespesa
        
        centro = CentroCusto.query.get_or_404(id)
        
        # Filtros de data
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        
        if data_inicio:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        if data_fim:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        
        # Buscar rateios
        rateios = RateioDespesa.get_por_centro_custo(id, data_inicio, data_fim)
        
        # Calcular total
        total = RateioDespesa.calcular_total_centro(id, data_inicio, data_fim)
        
        return render_template('financeiro/rateio/listar.html',
                             centro=centro,
                             rateios=rateios,
                             total=total,
                             data_inicio=data_inicio,
                             data_fim=data_fim)
    
    except Exception as e:
        flash(f'Erro: {str(e)}', 'danger')
        return redirect(url_for('financeiro.centros_custo'))


# ============================================================
# RELATÓRIOS CUSTOMIZÁVEIS
# ============================================================

@bp_financeiro.route('/relatorios-customizados')
def relatorios_customizados():
    """Lista relatórios customizados."""
    try:
        from app.financeiro.financeiro_model import RelatorioCustomizado
        
        relatorios = RelatorioCustomizado.query.filter_by(ativo=True).order_by(
            RelatorioCustomizado.favorito.desc(),
            RelatorioCustomizado.nome.asc()
        ).all()
        
        return render_template('financeiro/relatorios_customizados/listar.html',
                             relatorios=relatorios)
    
    except Exception as e:
        flash(f'Erro: {str(e)}', 'danger')
        return redirect(url_for('financeiro.dashboard'))


@bp_financeiro.route('/relatorios-customizados/novo', methods=['GET', 'POST'])
def novo_relatorio_customizado():
    """Criar novo relatório customizado."""
    if request.method == 'POST':
        try:
            from app.financeiro.financeiro_model import RelatorioCustomizado
            import json
            
            # Obter dados do formulário
            nome = request.form.get('nome')
            descricao = request.form.get('descricao')
            tipo = request.form.get('tipo')
            
            # Campos selecionados (checkboxes)
            campos = request.form.getlist('campos')
            
            # Filtros
            filtros = {}
            if request.form.get('filtro_tipo'):
                filtros['tipo'] = request.form.get('filtro_tipo')
            if request.form.get('filtro_status'):
                filtros['status'] = request.form.get('filtro_status')
            if request.form.get('filtro_data_inicio'):
                filtros['data_inicio'] = request.form.get('filtro_data_inicio')
            if request.form.get('filtro_data_fim'):
                filtros['data_fim'] = request.form.get('filtro_data_fim')
            
            # Criar relatório
            relatorio = RelatorioCustomizado(
                nome=nome,
                descricao=descricao,
                tipo=tipo,
                agrupamento=request.form.get('agrupamento'),
                ordenacao=request.form.get('ordenacao'),
                ordem_direcao=request.form.get('ordem_direcao', 'ASC'),
                formato_padrao=request.form.get('formato_padrao', 'EXCEL')
            )
            
            relatorio.set_campos(campos)
            relatorio.set_filtros(filtros)
            
            db.session.add(relatorio)
            db.session.commit()
            
            flash('Relatório criado com sucesso!', 'success')
            return redirect(url_for('financeiro.relatorios_customizados'))
        
        except Exception as e:
            flash(f'Erro: {str(e)}', 'danger')
            return redirect(request.url)
    
    # GET - Mostrar formulário
    from app.financeiro.financeiro_model import RelatorioCustomizado
    
    # Tipos disponíveis
    tipos = [
        {'valor': 'LANCAMENTOS', 'label': 'Lançamentos Financeiros'},
        {'valor': 'CONTAS_PAGAR', 'label': 'Contas a Pagar'},
        {'valor': 'CONTAS_RECEBER', 'label': 'Contas a Receber'},
        {'valor': 'FLUXO_CAIXA', 'label': 'Fluxo de Caixa'},
        {'valor': 'DRE', 'label': 'DRE'},
        {'valor': 'CENTROS_CUSTO', 'label': 'Centros de Custo'},
    ]
    
    return render_template('financeiro/relatorios_customizados/form.html',
                         tipos=tipos,
                         RelatorioCustomizado=RelatorioCustomizado)


@bp_financeiro.route('/relatorios-customizados/<int:id>/executar')
def executar_relatorio_customizado(id):
    """Executa relatório customizado."""
    try:
        from app.financeiro.financeiro_model import RelatorioCustomizado
        
        relatorio = RelatorioCustomizado.query.get_or_404(id)
        
        # Executar relatório
        dados = relatorio.executar()
        
        return render_template('financeiro/relatorios_customizados/resultado.html',
                             relatorio=relatorio,
                             dados=dados)
    
    except Exception as e:
        flash(f'Erro ao executar relatório: {str(e)}', 'danger')
        return redirect(url_for('financeiro.relatorios_customizados'))


@bp_financeiro.route('/relatorios-customizados/<int:id>/exportar/<formato>')
def exportar_relatorio_customizado(id, formato):
    """Exporta relatório customizado."""
    try:
        from app.financeiro.financeiro_model import RelatorioCustomizado
        from io import BytesIO
        import openpyxl
        
        relatorio = RelatorioCustomizado.query.get_or_404(id)
        dados = relatorio.executar()
        
        if formato == 'excel':
            # Criar Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = relatorio.nome[:31]  # Limite de 31 caracteres
            
            # Cabeçalho
            ws.append([relatorio.nome])
            ws.append([])
            
            # Dados
            if len(dados) > 0:
                # Obter campos
                campos = relatorio.get_campos_lista()
                
                # Cabeçalhos das colunas
                ws.append(campos)
                
                # Linhas de dados
                for item in dados:
                    linha = []
                    for campo in campos:
                        valor = getattr(item, campo, '')
                        linha.append(str(valor))
                    ws.append(linha)
            
            # Salvar
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f'{relatorio.nome.replace(" ", "_")}.xlsx'
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        
        else:
            flash('Formato não suportado', 'danger')
            return redirect(url_for('financeiro.relatorios_customizados'))
    
    except Exception as e:
        flash(f'Erro ao exportar: {str(e)}', 'danger')
        return redirect(url_for('financeiro.relatorios_customizados'))
