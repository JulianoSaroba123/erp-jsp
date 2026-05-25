# -*- coding: utf-8 -*-
"""
Rotas para Exportação e Importação de Dados - Módulo Energia Solar
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_from_directory
from flask_login import login_required, current_user
from app.extensoes import db
from app.energia_solar.energia_solar_model import CalculoEnergiaSolar
from app.energia_solar.catalogo_model import PlacaSolar, InversorSolar, KitSolar, ProjetoSolar
from app.energia_solar.custo_fixo_model import CustoPadraoSolar
from app.energia_solar.orcamento_model import OrcamentoItem
from datetime import datetime
from decimal import Decimal
import os
import json
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

exportacao_bp = Blueprint('exportacao', __name__, url_prefix='/energia-solar/exportacao',
                         template_folder='templates')


@exportacao_bp.route('/')
@login_required
def index():
    """Página principal de exportação/importação de dados"""
    try:
        # Listar arquivos de exportação existentes
        exports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'exports')
        arquivos_export = []
        
        if os.path.exists(exports_dir):
            for filename in os.listdir(exports_dir):
                if filename.startswith('energia_solar_export_') and (filename.endswith('.json') or filename.endswith('.xlsx')):
                    filepath = os.path.join(exports_dir, filename)
                    stat = os.stat(filepath)
                    arquivos_export.append({
                        'nome': filename,
                        'tamanho': stat.st_size,
                        'tamanho_mb': round(stat.st_size / 1024 / 1024, 2),
                        'data_criacao': datetime.fromtimestamp(stat.st_ctime),
                        'pode_baixar': True,
                        'tipo': 'Excel' if filename.endswith('.xlsx') else 'JSON'
                    })
            
            # Ordenar por data (mais recente primeiro)
            arquivos_export.sort(key=lambda x: x['data_criacao'], reverse=True)
        
        # Estatísticas do banco
        stats = {
            'calculos': CalculoEnergiaSolar.query.count(),
            'projetos': ProjetoSolar.query.count(),
            'kits': KitSolar.query.count(),
            'placas': PlacaSolar.query.count(),
            'inversores': InversorSolar.query.count(),
            'custos': CustoPadraoSolar.query.count(),
        }
        
        return render_template('energia_solar/exportacao.html', 
                             arquivos=arquivos_export,
                             stats=stats)
    
    except Exception as e:
        logger.error(f"Erro ao acessar página de exportação: {str(e)}")
        flash(f'Erro ao carregar página de exportação: {str(e)}', 'error')
        return redirect(url_for('energia_solar.index'))


@exportacao_bp.route('/executar', methods=['POST'])
@login_required
def executar():
    """Executa exportação de dados"""
    try:
        logger.info("🔄 Iniciando exportação de dados...")
        
        # Função para converter para JSON
        def converter_para_json(obj):
            if isinstance(obj, Decimal):
                return float(obj)
            if isinstance(obj, datetime):
                return obj.isoformat()
            raise TypeError(f"Tipo {type(obj)} não serializável")
        
        dados_exportados = {
            'metadata': {
                'data_exportacao': datetime.now().isoformat(),
                'versao': '1.0',
                'usuario': current_user.nome if hasattr(current_user, 'nome') else 'Sistema'
            }
        }
        
        # 1. Cálculos
        calculos = CalculoEnergiaSolar.query.all()
        dados_exportados['calculos_energia_solar'] = [
            {k: v for k, v in vars(calc).items() if not k.startswith('_')}
            for calc in calculos
        ]
        
        # 2. Projetos
        projetos = ProjetoSolar.query.all()
        dados_exportados['projetos_solares'] = [
            {k: v for k, v in vars(proj).items() if not k.startswith('_')}
            for proj in projetos
        ]
        
        # 3. Kits
        kits = KitSolar.query.filter_by(ativo=True).all()
        dados_exportados['kits_solares'] = [kit.to_dict() for kit in kits]
        
        # 4. Placas
        placas = PlacaSolar.query.filter_by(ativo=True).all()
        dados_exportados['placas_solares'] = [placa.to_dict() for placa in placas]
        
        # 5. Inversores
        inversores = InversorSolar.query.filter_by(ativo=True).all()
        dados_exportados['inversores_solares'] = [inv.to_dict() for inv in inversores]
        
        # 6. Custos Padrão
        custos = CustoPadraoSolar.query.filter_by(ativo=True).all()
        dados_exportados['custos_padrao'] = [
            {k: v for k, v in vars(custo).items() if not k.startswith('_')}
            for custo in custos
        ]
        
        # 7. Itens de Orçamento
        itens = OrcamentoItem.query.all()
        dados_exportados['orcamento_itens'] = [item.to_dict() for item in itens]
        
        # Criar pasta exports se não existir
        exports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'exports')
        os.makedirs(exports_dir, exist_ok=True)
        
        # Nome do arquivo com timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'energia_solar_export_{timestamp}.json'
        filepath = os.path.join(exports_dir, filename)
        
        # Salvar JSON
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(dados_exportados, f, ensure_ascii=False, indent=2, default=converter_para_json)
        
        total_registros = sum([
            len(dados_exportados['calculos_energia_solar']),
            len(dados_exportados['projetos_solares']),
            len(dados_exportados['kits_solares']),
            len(dados_exportados['placas_solares']),
            len(dados_exportados['inversores_solares']),
            len(dados_exportados['custos_padrao']),
            len(dados_exportados['orcamento_itens'])
        ])
        
        logger.info(f"✅ Exportação concluída: {total_registros} registros")
        
        return jsonify({
            'success': True,
            'filename': filename,
            'total_registros': total_registros,
            'detalhes': {
                'calculos': len(dados_exportados['calculos_energia_solar']),
                'projetos': len(dados_exportados['projetos_solares']),
                'kits': len(dados_exportados['kits_solares']),
                'placas': len(dados_exportados['placas_solares']),
                'inversores': len(dados_exportados['inversores_solares']),
                'custos': len(dados_exportados['custos_padrao']),
                'itens': len(dados_exportados['orcamento_itens'])
            }
        })
    
    except Exception as e:
        logger.error(f"❌ Erro na exportação: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@exportacao_bp.route('/executar-excel', methods=['POST'])
@login_required
def executar_excel():
    """Executa exportação de dados em formato Excel"""
    try:
        logger.info("🔄 Iniciando exportação Excel...")
        
        # Criar workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove a aba padrão
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        
        def criar_aba(titulo, dados, colunas):
            """Cria uma aba no Excel com os dados"""
            ws = wb.create_sheet(title=titulo)
            
            # Cabeçalhos
            for col_num, coluna in enumerate(colunas, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = coluna
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            
            # Dados
            for row_num, item in enumerate(dados, 2):
                for col_num, coluna in enumerate(colunas, 1):
                    valor = item.get(coluna.lower().replace(' ', '_'), '')
                    if isinstance(valor, datetime):
                        valor = valor.strftime('%d/%m/%Y %H:%M')
                    elif isinstance(valor, Decimal):
                        valor = float(valor)
                    ws.cell(row=row_num, column=col_num, value=valor)
            
            # Ajustar largura das colunas
            for col_num, _ in enumerate(colunas, 1):
                ws.column_dimensions[get_column_letter(col_num)].width = 15
            
            return len(dados)
        
        # 1. Cálculos
        calculos = CalculoEnergiaSolar.query.all()
        calculos_data = [calc.to_dict() for calc in calculos]
        criar_aba('Cálculos', calculos_data, ['ID', 'Consumo Mensal KWh', 'Potência Sistema KWp', 'Qtd Placas'])
        
        # 2. Projetos
        projetos = ProjetoSolar.query.all()
        projetos_data = [p.to_dict() for p in projetos]
        criar_aba('Projetos', projetos_data, ['ID', 'Cliente ID', 'Nome Cliente', 'Potência KWp', 'Status', 'Custo Total'])
        
        # 3. Kits
        kits = KitSolar.query.filter_by(ativo=True).all()
        kits_data = [kit.to_dict() for kit in kits]
        criar_aba('Kits', kits_data, ['ID', 'Nome', 'Potência', 'Preço', 'Fabricante'])
        
        # 4. Placas
        placas = PlacaSolar.query.filter_by(ativo=True).all()
        placas_data = [placa.to_dict() for placa in placas]
        criar_aba('Placas', placas_data, ['ID', 'Modelo', 'Potência', 'Eficiência', 'Preço', 'Fabricante'])
        
        # 5. Inversores
        inversores = InversorSolar.query.filter_by(ativo=True).all()
        inversores_data = [inv.to_dict() for inv in inversores]
        criar_aba('Inversores', inversores_data, ['ID', 'Modelo', 'Potência', 'Tipo', 'Preço', 'Fabricante'])
        
        # 6. Custos Padrão
        custos = CustoPadraoSolar.query.filter_by(ativo=True).all()
        custos_data = [custo.to_dict() for custo in custos]
        criar_aba('Custos Padrão', custos_data, ['ID', 'Descrição', 'Valor', 'Tipo'])
        
        # 7. Itens de Orçamento
        itens = OrcamentoItem.query.all()
        itens_data = [item.to_dict() for item in itens]
        criar_aba('Orçamento', itens_data, ['ID', 'Descrição', 'Quantidade', 'Preço Unitário', 'Total'])
        
        # Criar pasta exports se não existir
        exports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'exports')
        os.makedirs(exports_dir, exist_ok=True)
        
        # Nome do arquivo com timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'energia_solar_export_{timestamp}.xlsx'
        filepath = os.path.join(exports_dir, filename)
        
        # Salvar Excel
        wb.save(filepath)
        
        total_registros = len(calculos) + len(projetos) + len(kits) + len(placas) + len(inversores) + len(custos) + len(itens)
        
        logger.info(f"✅ Exportação Excel concluída: {total_registros} registros")
        
        return jsonify({
            'success': True,
            'filename': filename,
            'total_registros': total_registros,
            'detalhes': {
                'calculos': len(calculos),
                'projetos': len(projetos),
                'kits': len(kits),
                'placas': len(placas),
                'inversores': len(inversores),
                'custos': len(custos),
                'itens': len(itens)
            }
        })
    
    except Exception as e:
        logger.error(f"❌ Erro na exportação Excel: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@exportacao_bp.route('/download/<filename>')
@login_required
def download(filename):
    """Download de arquivo de exportação"""
    try:
        # Validar nome do arquivo
        valid_extensions = ('.json', '.xlsx')
        if not filename.startswith('energia_solar_export_') or not filename.endswith(valid_extensions):
            flash('Arquivo inválido', 'error')
            return redirect(url_for('exportacao.index'))
        
        exports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'exports')
        return send_from_directory(exports_dir, filename, as_attachment=True)
    
    except Exception as e:
        logger.error(f"Erro ao baixar arquivo: {str(e)}")
        flash(f'Erro ao baixar arquivo: {str(e)}', 'error')
        return redirect(url_for('exportacao.index'))


@exportacao_bp.route('/importar', methods=['POST'])
@login_required
def importar():
    """Importa dados de arquivo JSON"""
    try:
        # Verificar se arquivo foi enviado
        if 'arquivo' not in request.files:
            return jsonify({'success': False, 'error': 'Nenhum arquivo enviado'}), 400
        
        arquivo = request.files['arquivo']
        if arquivo.filename == '':
            return jsonify({'success': False, 'error': 'Nenhum arquivo selecionado'}), 400
        
        # Validar extensão
        if not arquivo.filename.endswith('.json'):
            return jsonify({'success': False, 'error': 'Apenas arquivos JSON são permitidos'}), 400
        
        # Ler conteúdo
        conteudo = arquivo.read().decode('utf-8')
        dados = json.loads(conteudo)
        
        # Função para converter strings ISO para datetime
        def converter_de_json(valor):
            if isinstance(valor, str):
                try:
                    if 'T' in valor or '-' in valor:
                        return datetime.fromisoformat(valor.replace('Z', '+00:00'))
                except:
                    pass
            return valor
        
        limpar_antes = request.form.get('limpar_antes') == 'true'
        
        logger.info(f"🔄 Iniciando importação (limpar_antes={limpar_antes})...")
        
        # Limpar dados se solicitado
        if limpar_antes:
            OrcamentoItem.query.delete()
            CalculoEnergiaSolar.query.delete()
            ProjetoSolar.query.delete()
            KitSolar.query.delete()
            PlacaSolar.query.delete()
            InversorSolar.query.delete()
            CustoPadraoSolar.query.delete()
            db.session.commit()
            logger.info("🗑️  Dados limpos antes da importação")
        
        estatisticas = {
            'placas': 0,
            'inversores': 0,
            'kits': 0,
            'projetos': 0,
            'calculos': 0,
            'custos': 0,
            'itens': 0
        }
        
        # Importar Placas
        for placa_data in dados.get('placas_solares', []):
            placa_id = placa_data.get('id')
            if not limpar_antes and PlacaSolar.query.get(placa_id):
                continue
            placa = PlacaSolar()
            for key, value in placa_data.items():
                if hasattr(placa, key) and key != 'id':
                    setattr(placa, key, converter_de_json(value))
            db.session.add(placa)
            estatisticas['placas'] += 1
        
        # Importar Inversores
        for inv_data in dados.get('inversores_solares', []):
            inv_id = inv_data.get('id')
            if not limpar_antes and InversorSolar.query.get(inv_id):
                continue
            inversor = InversorSolar()
            for key, value in inv_data.items():
                if hasattr(inversor, key) and key != 'id':
                    setattr(inversor, key, converter_de_json(value))
            db.session.add(inversor)
            estatisticas['inversores'] += 1
        
        db.session.commit()
        
        # Importar Kits (após placas e inversores)
        for kit_data in dados.get('kits_solares', []):
            kit_id = kit_data.get('id')
            if not limpar_antes and KitSolar.query.get(kit_id):
                continue
            kit = KitSolar()
            for key, value in kit_data.items():
                if hasattr(kit, key) and key != 'id':
                    setattr(kit, key, converter_de_json(value))
            db.session.add(kit)
            estatisticas['kits'] += 1
        
        # Importar Projetos
        for proj_data in dados.get('projetos_solares', []):
            proj_id = proj_data.get('id')
            if not limpar_antes and ProjetoSolar.query.get(proj_id):
                continue
            projeto = ProjetoSolar()
            for key, value in proj_data.items():
                if hasattr(projeto, key) and key != 'id':
                    setattr(projeto, key, converter_de_json(value))
            db.session.add(projeto)
            estatisticas['projetos'] += 1
        
        # Importar Custos Padrão
        for custo_data in dados.get('custos_padrao', []):
            custo_id = custo_data.get('id')
            if not limpar_antes and CustoPadraoSolar.query.get(custo_id):
                continue
            custo = CustoPadraoSolar()
            for key, value in custo_data.items():
                if hasattr(custo, key) and key != 'id':
                    if key in ['quantidade', 'valor_unitario', 'lucro_percentual']:
                        value = Decimal(str(value)) if value is not None else None
                    setattr(custo, key, converter_de_json(value))
            db.session.add(custo)
            estatisticas['custos'] += 1
        
        db.session.commit()
        
        total = sum(estatisticas.values())
        logger.info(f"✅ Importação concluída: {total} registros")
        
        return jsonify({
            'success': True,
            'total_registros': total,
            'detalhes': estatisticas
        })
    
    except Exception as e:
        db.session.rollback()
        logger.error(f"❌ Erro na importação: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@exportacao_bp.route('/excluir/<filename>', methods=['POST'])
@login_required
def excluir(filename):
    """Exclui arquivo de exportação"""
    try:
        # Validar nome do arquivo
        if not filename.startswith('energia_solar_export_') or not filename.endswith('.json'):
            return jsonify({'success': False, 'error': 'Arquivo inválido'}), 400
        
        exports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'exports')
        filepath = os.path.join(exports_dir, filename)
        
        if os.path.exists(filepath):
            os.remove(filepath)
            logger.info(f"🗑️  Arquivo excluído: {filename}")
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Arquivo não encontrado'}), 404
    
    except Exception as e:
        logger.error(f"Erro ao excluir arquivo: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500
